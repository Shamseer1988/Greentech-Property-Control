"""Migration off the master workbook.

Most tests build a miniature workbook in the same shape as the real one,
so they are fast and say precisely which rule they are about. The last
few run against the operator's actual file when it is present — that is
the only way to know the parser survives contact with fifteen years of
hand-keeping.

The recurring theme: the sheet is often wrong in small ways, and the
migration has to prefer the evidence that is *harder*. Rent columns beat
the Expire column, and rows beat totals.
"""
import io
import os
from datetime import date

import pytest
from openpyxl import Workbook

from app.services import migration

REAL_WORKBOOK = r"C:\Apps\Greentech Camp Ctrl\1-GreenTech Master File 2026.xlsm"
MONTHS = [date(2026, 1, 1), date(2026, 2, 1), date(2026, 3, 1)]


def build_workbook(tmp_path, blocks, landlords=None, receivables=None):
    """A miniature New-2026 + LANDLOARD in the real layout."""
    wb = Workbook()
    ws = wb.active
    ws.title = "New-2026"

    row = 1
    for index, block in enumerate(blocks, start=1):
        ws.cell(row, 1, index)
        ws.cell(row + 1, 1, block["title"])
        header = row + 2
        for col, label in enumerate(
                ["Client", "LAND LOARD", "Expire ", "MODE", "Rooms ", "OP Dec-2025"], 1):
            ws.cell(header, col, label)
        for offset, month in enumerate(MONTHS):
            ws.cell(header, 7 + offset, month)

        row = header + 1
        for tenant in block["tenants"]:
            ws.cell(row, 1, tenant["name"])
            ws.cell(row, 3, tenant.get("expire"))
            ws.cell(row, 4, tenant.get("mode", "CASH"))
            ws.cell(row, 5, tenant.get("rooms"))
            ws.cell(row, 6, tenant.get("opening"))
            for offset, value in enumerate(tenant.get("months", [])):
                if value is not None:
                    ws.cell(row, 7 + offset, value)
            row += 1

        for label, rooms, values in (
            ("TOTAL", block.get("total_rooms"), block.get("total")),
            ("Rent Paid", None, block.get("rent_paid")),
            ("Sewage Removal And Cleaning", None, block.get("sewage")),
            ("Eelctricity and Water", None, block.get("electricity")),
            ("Profit Or Loss", None, block.get("profit")),
        ):
            ws.cell(row, 1, label)
            if rooms is not None:
                ws.cell(row, 5, rooms)
            for offset, value in enumerate(values or []):
                if value is not None:
                    ws.cell(row, 7 + offset, value)
            row += 1
        row += 1

    if receivables:
        ws.cell(row, 1, "Cancelled & Pending Receivable")
        header = row + 1
        for col, label in enumerate(
                ["Client", "LAND LOARD", "Expire ", "MODE", "Rooms ", "OP Dec-2025"], 1):
            ws.cell(header, col, label)
        row = header + 1
        for entry in receivables:
            ws.cell(row, 1, entry["name"])
            ws.cell(row, 2, entry.get("property_title"))
            ws.cell(row, 3, "CANCELLED")
            ws.cell(row, 4, "CASH")
            ws.cell(row, 6, entry["opening"])
            row += 1

    ll = wb.create_sheet("LANDLOARD")
    for col, label in enumerate(
            ["S.no", "LAND LOARD", "ENGLISH", "ARABIC", "STREET & ROOMS",
             "START DATE", "EXP DATE", "RENT", "ROOMS", "STORE",
             "SECURITY CHQ", "OP Bls"], 1):
        ll.cell(1, col, label)
    for r, entry in enumerate(landlords or [], start=2):
        ll.cell(r, 1, r - 1)
        ll.cell(r, 2, entry["title"])
        ll.cell(r, 3, entry["english"])
        ll.cell(r, 4, entry.get("arabic"))
        ll.cell(r, 6, entry.get("start"))
        ll.cell(r, 7, entry.get("expiry"))
        ll.cell(r, 8, entry.get("rent"))
        ll.cell(r, 9, entry.get("rooms"))
        ll.cell(r, 10, entry.get("stores"))

    path = tmp_path / "book.xlsx"
    wb.save(path)
    return str(path)


SIMPLE = {
    "title": "AL SAFA ENGENERING STREET 10 [6 ROOMS] + 1 STORE",
    "total_rooms": 6,
    "tenants": [
        {"name": "TENANT ONE", "expire": date(2026, 12, 31), "rooms": 1,
         "months": [1400, 1400, 1400]},
        {"name": "TENANT TWO", "expire": date(2026, 12, 31), "rooms": 2,
         "mode": "CHEQUE", "opening": 5000, "months": [2600, 2600, 2800]},
    ],
    "rent_paid": [3000, 3000, 3000],
    "sewage": [100, 100, 100],
    "electricity": [500, 500, 500],
    "profit": [400, 400, 600],
}

SIMPLE_LANDLORD = [{
    "title": "AL SAFA ENGENERING STREET 10 [6 ROOMS] + 1 STORE",
    "english": "AL SAFA ENGENERING", "arabic": "الصفا",
    "start": date(2026, 1, 1), "expiry": date(2028, 1, 30),
    "rent": 3000, "rooms": 6, "stores": 1,
}]


def _actor(app):
    from app.models import User
    return User.query.filter_by(username="admin").first()


# ------------------------------------------------------------- parsing

def test_a_block_is_read_into_tenants_and_totals(app, tmp_path):
    path = build_workbook(tmp_path, [SIMPLE], SIMPLE_LANDLORD)
    with app.app_context():
        parsed = migration.parse_workbook(path)

    assert parsed["ok"] is True
    assert parsed["summary"]["properties"] == 1
    assert parsed["summary"]["tenants"] == 2

    block = parsed["blocks"][0]
    assert block["total_rooms"] == 6
    assert block["rent_paid"]["2026-01-01"] == 3000
    assert block["expenses"]["SEWAGE"]["2026-02-01"] == 100
    assert block["expenses"]["ELECTRICITY_CAMP"]["2026-03-01"] == 500

    two = block["tenants"][1]
    assert two["payment_mode"] == "cheque"
    assert two["rooms"] == 2
    assert two["opening_balance"] == 5000
    assert two["monthly"]["2026-03-01"] == 2800


def test_the_landlord_sheet_is_joined_by_title(app, tmp_path):
    path = build_workbook(tmp_path, [SIMPLE], SIMPLE_LANDLORD)
    with app.app_context():
        parsed = migration.parse_workbook(path)
    assert parsed["blocks"][0]["landlord_match"] == "exact"
    assert parsed["landlords"][0]["english"] == "AL SAFA ENGENERING"
    assert parsed["landlords"][0]["arabic"] == "الصفا"


def test_a_near_miss_title_is_matched_but_flagged(app, tmp_path):
    landlord = [dict(SIMPLE_LANDLORD[0],
                     title="MUBARAK SAFA ENGENERING STREET 10 (NO AGGREMENT)")]
    path = build_workbook(tmp_path, [SIMPLE], landlord)
    with app.app_context():
        parsed = migration.parse_workbook(path)
    block = parsed["blocks"][0]
    assert block["landlord_match"] == "fuzzy"
    assert any("similarity" in p["detail"] for p in parsed["problems"]), \
        "a guessed match must be flagged for confirmation"


def test_an_impossible_date_is_reported_not_guessed(app, tmp_path):
    landlord = [dict(SIMPLE_LANDLORD[0], expiry="31/06/2026")]   # June has 30 days
    path = build_workbook(tmp_path, [SIMPLE], landlord)
    with app.app_context():
        parsed = migration.parse_workbook(path)
    assert any("31/06/2026" in p["detail"] for p in parsed["problems"])
    assert parsed["landlords"][0]["expiry"] is None


@pytest.mark.parametrize("raw,expected", [
    ("CASH", "cash"), ("CHEQUE", "cheque"), ("ONLINE", "online"),
    ("M.CHEQUE", "cheque"),            # monthly cheque
    ("ONLINE AG.CHQ", "online"),       # paid online, cheque on file
    ("BANK TRANSFER", "online"),
])
def test_payment_modes_are_read_from_free_text(raw, expected):
    mode, problem = migration._payment_mode(raw)
    assert mode == expected, f"{raw!r} should be {expected}"
    assert problem is None


def test_an_unreadable_mode_falls_back_and_says_so():
    mode, problem = migration._payment_mode("???")
    assert mode == "cash"
    assert "unknown payment mode" in problem


def test_the_company_tail_is_not_swallowed_by_the_last_block(app, tmp_path):
    """The roll-up under the blocks belongs to the company, not to
    whichever building happens to be last on the sheet."""
    path = build_workbook(tmp_path, [SIMPLE], SIMPLE_LANDLORD,
                          receivables=[{"name": "GONE AWAY", "opening": 9000,
                                        "property_title": SIMPLE["title"]}])
    with app.app_context():
        parsed = migration.parse_workbook(path)
    names = [t["name"] for t in parsed["blocks"][0]["tenants"]]
    assert names == ["TENANT ONE", "TENANT TWO"]
    assert parsed["receivables"][0]["name"] == "GONE AWAY"
    assert parsed["receivables"][0]["opening_balance"] == 9000


# ---------------------------------------------------------------- plan

def test_two_blocks_on_one_street_get_different_names(app, tmp_path):
    """Three landlords have buildings on street 38. Collapsing them into
    one property would double-let its rooms."""
    a = dict(SIMPLE, title="OWNER A STREET NO.38 [4 ROOMS]", total_rooms=4)
    b = dict(SIMPLE, title="OWNER B STREET NO.38 [4 ROOMS]", total_rooms=4)
    path = build_workbook(tmp_path, [a, b], [])
    with app.app_context():
        plan = migration.build_commit_plan(migration.parse_workbook(path))
    names = [x["property_name"] for x in plan["blocks"]]
    assert len(set(names)) == 2, f"both blocks were named {names}"
    assert names[0] == "ST-38"


def test_the_plan_builds_enough_rooms_for_the_rows(app, tmp_path):
    """The TOTAL is hand-kept and drifts; the tenancies are the evidence."""
    block = dict(SIMPLE, total_rooms=1)          # rows need 3
    path = build_workbook(tmp_path, [block], SIMPLE_LANDLORD)
    with app.app_context():
        plan = migration.build_commit_plan(migration.parse_workbook(path))
    assert plan["blocks"][0]["rooms"] >= 3
    assert any("TOTAL says" in p["detail"] for p in plan["problems"])


# -------------------------------------------------------------- commit

def _migrate(app, path):
    from app.extensions import db
    with app.app_context():
        parsed = migration.parse_workbook(path)
        plan = migration.build_commit_plan(parsed)
        result = migration.commit_plan(plan, actor=_actor(app))
        db.session.commit()
        return parsed, plan, result


def test_commit_creates_the_estate(app, tmp_path):
    path = build_workbook(tmp_path, [SIMPLE], SIMPLE_LANDLORD)
    _, _, result = _migrate(app, path)

    from app.models import Client, ClientContract, Landlord, Property, Unit
    with app.app_context():
        assert Landlord.query.count() == 1
        assert Landlord.query.first().name_ar == "الصفا"
        prop = Property.query.one()
        assert prop.name == "ST-10"
        assert Unit.query.filter_by(property_id=prop.id).count() >= 6
        assert Client.query.count() == 2
        assert ClientContract.query.count() == 2

        two = ClientContract.query.join(Client).filter(
            Client.name == "TENANT TWO").one()
        assert two.payment_mode == "cheque"
        assert float(two.opening_balance) == 5000
    assert result["created"]["contracts"] == 2


def test_a_changing_month_column_becomes_a_dated_amendment(app, tmp_path):
    """2,600 → 2,800 in March is a rent change, and must be recorded as
    one so the rent engine reproduces the sheet month by month."""
    path = build_workbook(tmp_path, [SIMPLE], SIMPLE_LANDLORD)
    _migrate(app, path)

    from app.models import Client, ClientContract
    with app.app_context():
        contract = ClientContract.query.join(Client).filter(
            Client.name == "TENANT TWO").one()
        changes = [a for a in contract.amendments
                   if a.amendment_type == "rent_change"]
        assert len(changes) == 1
        assert float(changes[0].new_rent) == 2800
        assert changes[0].effective_date == date(2026, 3, 1)


def test_a_stale_expiry_does_not_truncate_a_billed_tenancy(app, tmp_path):
    """One real row carries an expiry of Jan-2025 while the sheet goes on
    billing it all through 2026. Honouring that date ends the tenancy on
    its first day and loses every rial."""
    block = dict(SIMPLE, tenants=[
        {"name": "STALE EXPIRY", "expire": date(2025, 1, 31), "rooms": 1,
         "months": [21000, 21000, 21000]},
    ])
    path = build_workbook(tmp_path, [block], SIMPLE_LANDLORD)
    _migrate(app, path)

    from app.models import ClientContract, RentCharge
    with app.app_context():
        contract = ClientContract.query.one()
        assert contract.expiry_date >= date(2026, 3, 31)
        charges = RentCharge.query.filter_by(contract_id=contract.id).all()
        billed = {c.period_month.isoformat(): float(c.amount) for c in charges}
        assert billed.get("2026-03-01") == 21000, billed


def test_a_row_that_stops_paying_is_a_tenancy_that_ended(app, tmp_path):
    block = dict(SIMPLE, tenants=[
        {"name": "LEAVER", "expire": date(2026, 12, 31), "rooms": 1,
         "months": [1400, 1400, None]},
    ])
    path = build_workbook(tmp_path, [block], SIMPLE_LANDLORD)
    _migrate(app, path)

    from app.models import ClientContract, RentCharge
    with app.app_context():
        contract = ClientContract.query.one()
        assert contract.status == "cancelled"
        assert contract.cancellation_date == date(2026, 2, 28)
        months = {c.period_month.isoformat()
                  for c in RentCharge.query.filter_by(contract_id=contract.id).all()}
        assert "2026-03-01" not in months, "billed a month the sheet did not"


def test_a_tenancy_with_nowhere_to_go_gets_a_unit_rather_than_being_dropped(
        app, tmp_path):
    block = dict(SIMPLE, total_rooms=1, tenants=[
        {"name": "FIRST", "expire": date(2026, 12, 31), "rooms": 1,
         "months": [1000, 1000, 1000]},
        {"name": "SECOND", "expire": date(2026, 12, 31), "rooms": 1,
         "months": [2000, 2000, 2000]},
    ])
    path = build_workbook(tmp_path, [block], [])
    _, _, result = _migrate(app, path)

    from app.models import ClientContract
    with app.app_context():
        assert ClientContract.query.count() == 2, "a paying tenancy was dropped"
    assert result["created"]["contracts"] == 2


def test_the_migration_can_be_run_twice(app, tmp_path):
    """A half-finished migration must be safe to re-run."""
    path = build_workbook(tmp_path, [SIMPLE], SIMPLE_LANDLORD)
    _migrate(app, path)
    _, _, second = _migrate(app, path)

    assert second["created"]["contracts"] == 0
    assert second["created"]["properties"] == 0
    assert second["created"]["landlords"] == 0
    assert second["reused"]["contracts"] == 2

    from app.models import ClientContract, Property
    with app.app_context():
        assert Property.query.count() == 1
        assert ClientContract.query.count() == 2


def test_a_departed_tenant_keeps_their_debt_without_holding_a_room(app, tmp_path):
    path = build_workbook(tmp_path, [SIMPLE], SIMPLE_LANDLORD,
                          receivables=[{"name": "GONE AWAY", "opening": 9000,
                                        "property_title": SIMPLE["title"]}])
    _, _, result = _migrate(app, path)
    assert result["created"]["receivables"] == 1

    from app.models import Client, ClientContract, RentCharge
    with app.app_context():
        contract = ClientContract.query.join(Client).filter(
            Client.name == "GONE AWAY").one()
        assert contract.status == "cancelled"
        assert contract.start_date.year == 2025, \
            "an old debt must not compete for rooms in the migrated year"
        opening = RentCharge.query.filter_by(
            contract_id=contract.id, is_opening_balance=True).one()
        assert float(opening.amount) == 9000


# --------------------------------------------------------- reconciliation

def test_reconciliation_reports_a_clean_month(app, tmp_path):
    path = build_workbook(tmp_path, [SIMPLE], SIMPLE_LANDLORD)
    parsed, plan, _ = _migrate(app, path)
    with app.app_context():
        report = migration.reconcile(parsed, month=date(2026, 3, 1), plan=plan)

    paid = next(l for l in report["company"] if l["label"] == "Paid to landlords")
    assert paid["workbook"] == 3000 and paid["app"] == 3000 and paid["matches"]
    sewage = next(l for l in report["company"]
                  if l["label"] == "Sewage removal and cleaning")
    assert sewage["matches"] is True
    assert report["summary"]["differing"] == 0, report


def test_reconciliation_shows_the_difference_rather_than_just_failing(
        app, tmp_path):
    """A mismatch has to be legible — which line, and by how much."""
    block = dict(SIMPLE, rent_paid=[3000, 3000, 9999])
    path = build_workbook(tmp_path, [block], SIMPLE_LANDLORD)
    parsed, plan, _ = _migrate(app, path)

    from app.extensions import db
    from app.models import LandlordPayment
    with app.app_context():
        march = LandlordPayment.query.filter_by(
            period_month=date(2026, 3, 1)).one()
        march.amount = 4000                       # pretend somebody mistyped it
        db.session.commit()
        report = migration.reconcile(parsed, month=date(2026, 3, 1), plan=plan)

    paid = next(l for l in report["company"] if l["label"] == "Paid to landlords")
    assert paid["matches"] is False
    assert paid["workbook"] == 9999 and paid["app"] == 4000
    assert paid["difference"] == -5999


# ------------------------------------------------------------ the real file

real_only = pytest.mark.skipif(
    not os.path.exists(REAL_WORKBOOK),
    reason="the operator's master workbook is not on this machine")


@real_only
def test_the_real_workbook_parses_without_blockers(app):
    with app.app_context():
        parsed = migration.parse_workbook(REAL_WORKBOOK)
    assert parsed["ok"] is True, parsed["problems"]
    assert parsed["summary"]["properties"] == 18
    assert parsed["summary"]["tenants"] > 100
    assert parsed["summary"]["blockers"] == 0


@real_only
def test_the_real_workbook_reconciles_for_june(app):
    """The Phase-8 exit criterion, against the operator's own file: a
    month closes in the app alone and every figure ties back to Excel."""
    from app.extensions import db
    with app.app_context():
        parsed = migration.parse_workbook(REAL_WORKBOOK)
        # Row 185 in the ST-38 HYPER block ("MIHRAB ROOM RENT ADJUST
        # (1000*2)") is a manual note, not a tenant -- it carries a
        # negative one-off figure that create_contract rightly refuses.
        # The operator's call: keep MIHRAB-STAFF's rent at its normal
        # 5,000 and book the 4,000 as a separate property expense
        # instead (done once, by hand, outside this migration).
        plan = migration.build_commit_plan(
            parsed, overrides={"skip_tenants": {"7": [185]}})
        migration.commit_plan(plan, actor=_actor(app))
        db.session.commit()
        report = migration.reconcile(parsed, month=date(2026, 6, 1), plan=plan)

    assert report["summary"]["differing"] == 0, [
        l for l in report["company"] + report["properties"]
        if l["matches"] is False]
    assert report["summary"]["matching"] >= 20

    charged = next(l for l in report["company"]
                   if l["label"] == "Rent charged to clients")
    assert charged["workbook"] == 882813.00
    assert charged["app"] == 882813.00
