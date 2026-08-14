"""Dashboards and the reports that replace the workbook's sheets.

The recurring assertion here is agreement: a figure on a dashboard and
the same figure on its report come from one service, so they must match
to the cent. A report that quietly disagrees with the screen is worse
than no report.
"""
import io
from datetime import date, timedelta

import pytest
from openpyxl import load_workbook


MONTH = "2026-03-01"


@pytest.fixture()
def estate(client, auth_headers):
    """A property let to two clients: one pays, one doesn't.

    ST-40 has 4 units. IMDAAD holds 2 at 10,000; NAKHEEL holds 1 at
    6,000; one unit stays empty. March rent is raised for both, IMDAAD
    pays in full, NAKHEEL pays nothing — so every report has both a
    settled and an outstanding row to show.
    """
    landlord = client.post("/api/v1/landlords", headers=auth_headers,
                           json={"name": "AL SAFA ENGENERING"}).get_json()["data"]
    prop = client.post("/api/v1/properties", headers=auth_headers, json={
        "name": "ST-40", "property_type": "building_with_store",
        "landlord_id": landlord["id"],
        "layout": {"floors": 1, "units_per_floor": 4},
    }).get_json()["data"]
    client.post(f"/api/v1/properties/{prop['id']}/agreements", headers=auth_headers,
                json={"landlord_id": landlord["id"], "start_date": "2026-01-01",
                      "expiry_date": "2026-12-31", "monthly_rent": 12000,
                      "security_deposit": 12000})
    units = client.get(f"/api/v1/properties/{prop['id']}/units",
                       headers=auth_headers).get_json()["data"]

    payers = {}
    for name, unit_ids, rent, mode in (
        ("IMDAAD FACILITY SERVICES", [units[0]["id"], units[1]["id"]], 10000, "cash"),
        ("NAKHEEL LANDSCAPES", [units[2]["id"]], 6000, "cheque"),
    ):
        tenant = client.post("/api/v1/clients", headers=auth_headers,
                             json={"name": name}).get_json()["data"]
        contract = client.post("/api/v1/contracts", headers=auth_headers, json={
            "client_id": tenant["id"], "property_id": prop["id"],
            "unit_ids": unit_ids, "start_date": "2026-01-01",
            "expiry_date": "2026-12-31", "monthly_rent": rent,
            "payment_mode": mode,
        }).get_json()["data"]
        client.post("/api/v1/rent/generate", headers=auth_headers,
                    json={"contract_id": contract["id"], "upto": MONTH})
        payers[name] = {"client": tenant, "contract": contract}

    # IMDAAD settles March; NAKHEEL does not.
    client.post("/api/v1/rent/receipts", headers=auth_headers, json={
        "client_id": payers["IMDAAD FACILITY SERVICES"]["client"]["id"],
        "amount": 30000, "receipt_date": "2026-03-05", "mode": "cash",
    })
    client.post("/api/v1/expenses/landlord-payments", headers=auth_headers, json={
        "landlord_id": landlord["id"], "property_id": prop["id"],
        "period_month": MONTH, "amount": 12000, "mode": "cheque",
    })
    return {"landlord": landlord, "property": prop, "units": units, **payers}


def _report(client, auth_headers, slug, **params):
    query = "&".join(f"{k}={v}" for k, v in params.items())
    resp = client.get(f"/api/v1/reports/{slug}?{query}", headers=auth_headers)
    assert resp.status_code == 200, resp.get_data(as_text=True)
    return resp.get_json()["data"]


# ------------------------------------------------------------- catalog

def test_every_workbook_sheet_has_a_report(client, auth_headers):
    """The Phase-6 exit criterion, as a list."""
    slugs = {r["slug"] for r in
             client.get("/api/v1/reports", headers=auth_headers).get_json()["data"]}
    for required in ("rent-collection", "ageing", "empty-units", "contract-expiry",
                     "property-pnl", "company-pnl", "pdc-register", "audit-trail"):
        assert required in slugs, f"{required} is missing from the catalog"


# ---------------------------------------------------------- main dashboard

def test_main_dashboard_reports_the_month(client, auth_headers, estate):
    data = client.get(f"/api/v1/dashboard/summary?month={MONTH}",
                      headers=auth_headers).get_json()["data"]

    assert data["month"] == MONTH
    assert data["properties"]["total"] == 1
    assert data["clients"]["total"] == 2
    assert data["contracts"]["active"] == 2

    # 3 of 4 units let.
    assert data["units"]["occupied"] == 3
    assert data["units"]["empty"] == 1

    # 16,000 charged for March, 10,000 of it settled.
    assert data["collections"]["charged"] == 16000.0
    assert data["collections"]["collected"] == 10000.0
    assert data["collections"]["outstanding"] == 6000.0
    assert data["collections"]["collection_percent"] == 62.5


def test_dashboard_and_reports_cannot_disagree(client, auth_headers, estate):
    """Same number, two surfaces — this is the whole point of routing
    both through one service."""
    dash = client.get(f"/api/v1/dashboard/summary?month={MONTH}",
                      headers=auth_headers).get_json()["data"]
    collection = _report(client, auth_headers, "rent-collection", month=MONTH)
    ageing = _report(client, auth_headers, "ageing", upto=MONTH)
    pnl = _report(client, auth_headers, "property-pnl", month=MONTH)

    assert collection["meta"]["charged"] == dash["collections"]["charged"]
    assert collection["meta"]["collected"] == dash["collections"]["collected"]
    assert ageing["meta"]["grand_total"] == dash["ageing"]["total_outstanding"]
    assert pnl["meta"]["net_profit"] == dash["pnl"]["net_profit"]


def test_dashboard_flags_what_needs_attention(client, auth_headers, estate):
    data = client.get("/api/v1/dashboard/summary", headers=auth_headers).get_json()["data"]
    assert "buckets" in data["contract_expiry"]
    assert "due_this_week" in data["cheques"]
    assert "bounced" in data["cheques"]
    assert data["approvals"]["total"] == 0


def test_dashboard_rejects_a_bad_month(client, auth_headers):
    resp = client.get("/api/v1/dashboard/summary?month=March", headers=auth_headers)
    assert resp.status_code == 400


# -------------------------------------------------------- entity dashboards

def test_property_dashboard_is_the_new_2026_block(client, auth_headers, estate):
    pid = estate["property"]["id"]
    data = client.get(f"/api/v1/dashboard/property/{pid}?month={MONTH}",
                      headers=auth_headers).get_json()["data"]

    assert data["property"]["name"] == "ST-40"
    assert data["landlord"]["name"] == "AL SAFA ENGENERING"
    assert data["agreement"]["monthly_rent"] == 12000.0
    assert data["units"]["total"] == 4 and data["units"]["empty"] == 1
    assert data["units"]["occupancy_percent"] == 75.0

    assert len(data["contracts"]) == 2
    imdaad = next(c for c in data["contracts"] if c["client_name"].startswith("IMDAAD"))
    assert sorted(imdaad["units"]) == ["101", "102"]

    # 16,000 charged − 12,000 to the landlord.
    assert data["pnl"]["rent_charged"] == 16000.0
    assert data["pnl"]["rent_paid"] == 12000.0
    assert data["pnl"]["profit"] == 4000.0

    assert len(data["trend"]) == 12
    assert data["trend"][-1]["month"] == MONTH
    assert data["trend"][-1]["profit"] == 4000.0


def test_client_dashboard_shows_what_they_hold_and_owe(client, auth_headers, estate):
    cid = estate["NAKHEEL LANDSCAPES"]["client"]["id"]
    data = client.get(f"/api/v1/dashboard/client/{cid}?month={MONTH}",
                      headers=auth_headers).get_json()["data"]

    assert data["client"]["name"] == "NAKHEEL LANDSCAPES"
    assert data["active_contracts"] == 1
    assert data["contracts"][0]["units"] == ["103"]
    assert data["outstanding"] == 18000.0        # Jan + Feb + Mar at 6,000
    assert data["recent_receipts"] == []


def test_client_dashboard_for_a_payer_shows_the_receipt(client, auth_headers, estate):
    cid = estate["IMDAAD FACILITY SERVICES"]["client"]["id"]
    data = client.get(f"/api/v1/dashboard/client/{cid}?month={MONTH}",
                      headers=auth_headers).get_json()["data"]
    assert len(data["recent_receipts"]) == 1
    assert data["recent_receipts"][0]["amount"] == 30000.0


def test_client_dashboard_lists_the_cheques_still_on_hand(client, auth_headers, estate):
    """A freshly registered PDC has status "received" — not "pending",
    which is not a cheque status at all. Getting that wrong silently
    shows a client's cheque book as empty."""
    contract = estate["NAKHEEL LANDSCAPES"]["contract"]
    client.post(f"/api/v1/contracts/{contract['id']}/cheques", headers=auth_headers,
                json={"cheques": [{"cheque_number": "000501",
                                   "cheque_date": "2026-01-05", "amount": 6000,
                                   "bank_name": "QNB", "for_month": "2026-01-01"}],
                      "security": {"cheque_number": "SEC-01",
                                   "cheque_date": "2026-12-31", "amount": 6000}})

    cid = estate["NAKHEEL LANDSCAPES"]["client"]["id"]
    data = client.get(f"/api/v1/dashboard/client/{cid}", headers=auth_headers).get_json()["data"]
    assert len(data["cheques"]) == 2, "cheques on hand are missing from the client page"
    assert {c["status"] for c in data["cheques"]} == {"received"}


def test_landlord_dashboard_shows_commitment_and_payments(client, auth_headers, estate):
    lid = estate["landlord"]["id"]
    data = client.get(f"/api/v1/dashboard/landlord/{lid}?month={MONTH}",
                      headers=auth_headers).get_json()["data"]

    assert data["landlord"]["name"] == "AL SAFA ENGENERING"
    assert len(data["properties"]) == 1
    assert data["monthly_commitment"] == 12000.0
    assert data["total_paid_to_date"] == 12000.0
    assert data["agreements"][0]["security_deposit"] == 12000.0
    assert data["per_property"][0]["profit"] == 4000.0


def test_entity_dashboards_404_on_a_missing_record(client, auth_headers):
    for path in ("property/9999", "client/9999", "landlord/9999"):
        assert client.get(f"/api/v1/dashboard/{path}",
                          headers=auth_headers).status_code == 404


# ------------------------------------------------------------- reports

def test_rent_collection_lists_both_clients(client, auth_headers, estate):
    data = _report(client, auth_headers, "rent-collection", month=MONTH)
    assert data["meta"]["count"] == 2
    paid = next(r for r in data["rows"] if r["client_name"].startswith("IMDAAD"))
    unpaid = next(r for r in data["rows"] if r["client_name"].startswith("NAKHEEL"))
    assert paid["collected"] == 10000.0 and paid["outstanding"] == 0.0
    assert unpaid["collected"] == 0.0 and unpaid["outstanding"] == 6000.0


def test_ageing_puts_a_month_in_every_column(client, auth_headers, estate):
    data = _report(client, auth_headers, "ageing", upto=MONTH, months=6)
    assert data["meta"]["grand_total"] == 18000.0
    row = data["rows"][0]
    assert row["client_name"] == "NAKHEEL LANDSCAPES"
    assert row["2026-01-01"] == 6000.0
    assert row["2026-03-01"] == 6000.0
    assert row["total"] == 18000.0
    assert any(c["key"] == "2026-02-01" for c in data["columns"])


def test_empty_units_names_the_unit_number(client, auth_headers, estate):
    data = _report(client, auth_headers, "empty-units")
    assert data["meta"]["count"] == 1
    assert data["rows"][0]["unit_number"] == "104"


def test_contract_expiry_buckets_by_days_left(client, auth_headers, estate):
    data = _report(client, auth_headers, "contract-expiry", all="1")
    assert data["meta"]["count"] == 2
    assert {r["bucket"] for r in data["rows"]} <= {
        "expired", "30 days", "60 days", "90 days"}


def test_property_pnl_makes_a_column_per_expense_category(client, auth_headers, estate):
    data = _report(client, auth_headers, "property-pnl", month=MONTH)
    assert data["meta"]["count"] == 1
    row = data["rows"][0]
    assert row["property_name"] == "ST-40"
    assert row["rent_charged"] == 16000.0
    assert row["rent_paid"] == 12000.0
    assert row["profit"] == 4000.0


def test_company_pnl_starts_at_income_and_ends_at_net(client, auth_headers, estate):
    data = _report(client, auth_headers, "company-pnl", month=MONTH)
    assert data["rows"][0]["line"] == "Rent charged to clients"
    assert data["rows"][0]["amount"] == 16000.0
    assert data["rows"][-1]["line"] == "Net profit"
    assert data["rows"][-1]["amount"] == data["meta"]["net_profit"]
    # Rent paid to the landlord is shown as a cost, hence negative.
    paid = next(r for r in data["rows"] if r["line"] == "Rent paid to landlords")
    assert paid["amount"] == -12000.0


def test_company_pnl_shows_income_kind_categories_as_income_not_negated_overhead(
        client, auth_headers, estate):
    """Regression: an income-kind category (e.g. Other Income) used to
    fall into the `else` branch of the section bucketing and get
    negated into Overhead. It must show as a positive Income line."""
    categories = client.get("/api/v1/expenses/categories", headers=auth_headers).get_json()["data"]
    other_income = next(c for c in categories if c["code"] == "OTHER_INCOME")
    client.post("/api/v1/expenses", headers=auth_headers, json={
        "category_id": other_income["id"], "period_month": MONTH, "amount": 1500,
    })
    data = _report(client, auth_headers, "company-pnl", month=MONTH)
    row = next(r for r in data["rows"] if r["line"] == "Other Income")
    assert row["section"] == "Income"
    assert row["amount"] == 1500.0, "must not be negated"


def test_company_pnl_excludes_rent_paid_category_to_avoid_double_count(
        client, auth_headers, estate):
    """RENT_PAID is already reflected via "Rent paid to landlords"
    (sourced from LandlordPayment); an Expense row under that category
    must not also appear as its own line."""
    categories = client.get("/api/v1/expenses/categories", headers=auth_headers).get_json()["data"]
    rent_paid_cat = next(c for c in categories if c["code"] == "RENT_PAID")
    client.post("/api/v1/expenses", headers=auth_headers, json={
        "category_id": rent_paid_cat["id"], "property_id": estate["property"]["id"],
        "period_month": MONTH, "amount": 9999,
    })
    data = _report(client, auth_headers, "company-pnl", month=MONTH)
    names = [r["line"] for r in data["rows"]]
    assert names.count("Rent Paid") == 0, "RENT_PAID category must not appear as its own line"
    paid = next(r for r in data["rows"] if r["line"] == "Rent paid to landlords")
    assert paid["amount"] == -12000.0, "must stay the LandlordPayment figure, not +9999 more"


def test_monthly_pnl_reproduces_the_workbook_layout(client, auth_headers, estate):
    """Rows are P&L lines, columns are months plus a Total — and the
    figures must foot: Total Revenues - Cost of Sales = Gross Profit;
    Gross Profit - Total Indirect = Net Profit."""
    categories = {c["code"]: c for c in client.get(
        "/api/v1/expenses/categories", headers=auth_headers).get_json()["data"]}
    client.post("/api/v1/expenses", headers=auth_headers, json={
        "category_id": categories["OTHER_INCOME"]["id"], "period_month": MONTH, "amount": 1500,
    })
    client.post("/api/v1/expenses", headers=auth_headers, json={
        "category_id": categories["SALARY"]["id"], "period_month": MONTH, "amount": 65900,
    })

    data = _report(client, auth_headers, "monthly-pnl", from_month=MONTH, to_month=MONTH)
    month_key = MONTH[:7]
    by_line = {r["line"]: r for r in data["rows"]}

    assert by_line["Other Income"][month_key] == 1500.0
    assert by_line["Rent Received"][month_key] == 16000.0
    assert by_line["Total Revenues"][month_key] == 17500.0
    assert by_line["Rent Paid"][month_key] == 12000.0, "accrual, falls back to cash figure"
    assert by_line["Cost of Sales"][month_key] == 12000.0
    assert by_line["Gross Profit / (Loss)"][month_key] == 17500.0 - 12000.0
    assert by_line["Salary and Allowances"][month_key] == 65900.0
    assert by_line["Total Indirect Expenses"][month_key] == 65900.0
    expected_net = (17500.0 - 12000.0) - 65900.0
    assert by_line["Net Profit / (Loss)"][month_key] == expected_net
    assert by_line["Net Profit / (Loss)"]["total"] == expected_net, "single-month range: total = the month"


def test_pdc_register_lists_the_cheque_book(client, auth_headers, estate):
    contract = estate["NAKHEEL LANDSCAPES"]["contract"]
    resp = client.post(f"/api/v1/contracts/{contract['id']}/cheques", headers=auth_headers,
                       json={
                           "cheques": [
                               {"cheque_number": "000501", "cheque_date": "2026-01-05",
                                "amount": 6000, "bank_name": "QNB",
                                "for_month": "2026-01-01"},
                               {"cheque_number": "000502", "cheque_date": "2026-02-05",
                                "amount": 6000, "bank_name": "QNB",
                                "for_month": "2026-02-01"},
                           ],
                           "security": {"cheque_number": "SEC-01",
                                        "cheque_date": "2026-12-31",
                                        "amount": 6000, "bank_name": "QNB"},
                       })
    assert resp.status_code in (200, 201), resp.get_data(as_text=True)
    data = _report(client, auth_headers, "pdc-register")
    assert data["meta"]["count"] == 3
    assert data["meta"]["total"] == 18000.0
    assert sum(1 for r in data["rows"] if r["kind"] == "Security") == 1


def test_audit_report_records_the_work_just_done(client, auth_headers, estate):
    data = _report(client, auth_headers, "audit-trail", module="contract")
    assert data["meta"]["count"] >= 1
    assert {"create"} <= {r["action"] for r in data["rows"]}


# -------------------------------------------------------------- exports

@pytest.mark.parametrize("slug", [
    "rent-collection", "ageing", "empty-units", "contract-expiry",
    "property-pnl", "company-pnl", "pdc-register", "audit-trail",
])
def test_every_report_exports_to_excel_and_pdf(client, auth_headers, estate, slug):
    xlsx = client.get(f"/api/v1/reports/{slug}/export?month={MONTH}",
                      headers=auth_headers)
    assert xlsx.status_code == 200, xlsx.get_data(as_text=True)[:300]
    assert xlsx.data[:2] == b"PK", "not a zip, so not a workbook"
    assert slug in xlsx.headers["Content-Disposition"]

    pdf = client.get(f"/api/v1/reports/{slug}/export?format=pdf&month={MONTH}",
                     headers=auth_headers)
    assert pdf.status_code == 200, pdf.get_data(as_text=True)[:300]
    assert pdf.data[:4] == b"%PDF"
    assert pdf.mimetype == "application/pdf"


def test_the_excel_export_carries_the_real_figures(client, auth_headers, estate):
    resp = client.get("/api/v1/reports/rent-collection/export?month=" + MONTH,
                      headers=auth_headers)
    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert "Charged" in header and "Outstanding" in header

    charged_col = header.index("Charged") + 1
    values = sorted(ws.cell(r, charged_col).value
                    for r in range(2, ws.max_row + 1))
    assert values == [6000.0, 10000.0]


def test_an_unknown_export_format_is_refused(client, auth_headers):
    resp = client.get("/api/v1/reports/ageing/export?format=docx", headers=auth_headers)
    assert resp.status_code == 400
    assert "pdf" in resp.get_json()["message"]


def test_export_needs_the_export_permission(client):
    assert client.get("/api/v1/reports/ageing/export").status_code == 401
