import io
from datetime import date, timedelta
from openpyxl import load_workbook


def _seed(client, auth_headers):
    """One property with an agreement expiring in 20 days and 2 units,
    one of them occupied."""
    today = date.today()
    landlord = client.post("/api/v1/landlords", headers=auth_headers,
                           json={"name": "Mansoor"}).get_json()["data"]
    prop = client.post("/api/v1/properties", headers=auth_headers,
                       json={"name": "Doha B12", "property_type": "full_building",
                             "city": "Doha", "landlord_id": landlord["id"]}).get_json()["data"]
    client.post(f"/api/v1/properties/{prop['id']}/agreements", headers=auth_headers,
                json={
                    "landlord_id": landlord["id"],
                    "start_date": (today - timedelta(days=200)).isoformat(),
                    "expiry_date": (today + timedelta(days=20)).isoformat(),
                    "monthly_rent": 10000,
                })
    floor = client.post(f"/api/v1/properties/{prop['id']}/floors", headers=auth_headers,
                        json={"floor_number": "1"}).get_json()["data"]
    r1 = client.post(f"/api/v1/floors/{floor['id']}/units", headers=auth_headers,
                     json={"unit_number": "101", "monthly_rent": 1400}).get_json()["data"]
    r2 = client.post(f"/api/v1/floors/{floor['id']}/units", headers=auth_headers,
                     json={"unit_number": "102", "monthly_rent": 1400}).get_json()["data"]
    client.post(f"/api/v1/units/{r1['id']}/status", headers=auth_headers,
                json={"status": "occupied"})
    return {"property": prop, "floor": floor, "units": [r1, r2], "landlord": landlord}


def test_list_reports_catalog(client, auth_headers):
    resp = client.get("/api/v1/reports", headers=auth_headers)
    assert resp.status_code == 200
    slugs = {r["slug"] for r in resp.get_json()["data"]}
    for must_have in ["property-occupancy", "empty-units", "agreement-expiry"]:
        assert must_have in slugs


def test_property_occupancy_report(client, auth_headers):
    s = _seed(client, auth_headers)
    resp = client.get("/api/v1/reports/property-occupancy", headers=auth_headers)
    assert resp.status_code == 200
    payload = resp.get_json()["data"]
    assert "columns" in payload and "rows" in payload
    row = next(r for r in payload["rows"] if r["code"] == s["property"]["code"])
    assert row["total"] == 2
    assert row["occupied"] == 1
    assert row["empty"] == 1
    assert row["occupancy_percent"] == 50.0


def test_empty_units_report(client, auth_headers):
    s = _seed(client, auth_headers)
    resp = client.get(f"/api/v1/reports/empty-units?property_id={s['property']['id']}",
                      headers=auth_headers)
    payload = resp.get_json()["data"]
    assert payload["meta"]["count"] == 1
    row = payload["rows"][0]
    assert row["unit_number"] == "102"
    assert row["property_code"] == s["property"]["code"]
    assert row["monthly_rent"] == 1400


def test_agreement_expiry_report(client, auth_headers):
    s = _seed(client, auth_headers)
    resp = client.get("/api/v1/reports/agreement-expiry", headers=auth_headers)
    payload = resp.get_json()["data"]
    rows = payload["rows"]
    assert len(rows) >= 1
    only = next(r for r in rows if r["property_code"] == s["property"]["code"])
    assert only["landlord"] == s["landlord"]["name"]
    assert only["bucket"] == "30"
    assert only["days_left"] == 20
    assert only["monthly_rent"] == 10000


def test_agreement_expiry_bucket_filter(client, auth_headers):
    _seed(client, auth_headers)
    resp = client.get("/api/v1/reports/agreement-expiry?bucket=7", headers=auth_headers)
    assert resp.get_json()["data"]["meta"]["count"] == 0


def test_excel_export_returns_xlsx(client, auth_headers):
    _seed(client, auth_headers)
    resp = client.get("/api/v1/reports/property-occupancy/export", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert resp.data[:2] == b"PK"
    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert "Property" in header
    assert "Occupancy %" in header


def test_unknown_report_404(client, auth_headers):
    resp = client.get("/api/v1/reports/totally-fake", headers=auth_headers)
    assert resp.status_code == 404


def test_report_view_requires_permission(client):
    resp = client.get("/api/v1/reports/property-occupancy")
    assert resp.status_code == 401
