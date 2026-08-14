"""Landlord payments, expenses, the P&L import, and property-wise P&L."""
import io
from datetime import date

from openpyxl import Workbook


# ----------------------------------------------------------------------
# A stand-in for the accounting software's export, built to the exact
# layout of the real file: title carrying the period, indented ledger
# names in column A, amounts in B, section totals in C.
# ----------------------------------------------------------------------

DEFAULT_REVENUE = [("RENT RECEIVED", 873813), ("OTHER INCOME", 1500)]
DEFAULT_DIRECT = [
    ("SEAWAGE REMOVAL AND CLEANING", 33845),
    ("ELECTRICITY AND WATER FOR CAMP", 139175),
    ("RENT PAID", 499885),
    ("GENERAL EXPENCE FOR CAMP", 4169.5),
]
DEFAULT_INDIRECT = [
    ("BANK CHARGES", 1656.4), ("SALARY AND ALLOWANCES", 65900),
    ("FUEL CHARGES", 3035), ("TELEPHONE CHARGES", 262.9),
]


def build_pnl(revenue=None, direct=None, indirect=None,
              period="from 01-Jun-2026 to 30-Jun-2026",
              break_totals=False) -> io.BytesIO:
    revenue = DEFAULT_REVENUE if revenue is None else revenue
    direct = DEFAULT_DIRECT if direct is None else direct
    indirect = DEFAULT_INDIRECT if indirect is None else indirect

    wb = Workbook()
    ws = wb.active
    ws.title = "RepTPL"
    ws["A2"] = "GREENTEC TRADING & CONTRACTING LLC"
    ws["A3"] = f"Trading And Profit & Loss Account {period} (Net Transaction Only)"
    ws["A5"], ws["B5"], ws["C5"] = "Particulars", "Amount", "Total"

    row = 6
    ws.cell(row, 1, "Revenues"); row += 1
    for name, amount in revenue:
        ws.cell(row, 1, "          " + name)
        ws.cell(row, 2, amount)
        row += 1
    total_rev = sum(a for _, a in revenue)
    ws.cell(row, 1, "Total Revenues")
    ws.cell(row, 3, total_rev + (1000 if break_totals else 0)); row += 1

    ws.cell(row, 1, "     Purchase & Direct Expenses"); row += 1
    for name, amount in direct:
        ws.cell(row, 1, "          " + name)
        ws.cell(row, 2, amount)
        row += 1
    total_direct = sum(a for _, a in direct)
    ws.cell(row, 1, "Cost of Sales"); ws.cell(row, 3, total_direct); row += 1
    ws.cell(row, 1, "(Gross Profit)"); ws.cell(row, 3, total_rev - total_direct); row += 2

    ws.cell(row, 1, "Indirect Expenses"); row += 1
    for name, amount in indirect:
        ws.cell(row, 1, "     " + name)
        ws.cell(row, 2, amount)
        row += 1
    total_indirect = sum(a for _, a in indirect)
    ws.cell(row, 1, "Total Indirect Expenses"); ws.cell(row, 3, total_indirect); row += 1
    ws.cell(row, 1, "(Net Profit)")
    ws.cell(row, 3, total_rev - total_direct - total_indirect)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _upload(client, auth_headers, stream, name="Profit and Loss.xlsx"):
    return client.post("/api/v1/expenses/import/preview", headers=auth_headers,
                       data={"file": (stream, name)},
                       content_type="multipart/form-data")


def _property(client, auth_headers, name="ST-10"):
    landlord = client.post("/api/v1/landlords", headers=auth_headers,
                           json={"name": f"{name} Owner"}).get_json()["data"]
    prop = client.post("/api/v1/properties", headers=auth_headers, json={
        "name": name, "property_type": "labour_camp", "landlord_id": landlord["id"],
        "layout": {"floors": 1, "units_per_floor": 3},
    }).get_json()["data"]
    return landlord, prop


# ---------------------------------------------------------------- parser

def test_preview_reads_period_lines_and_totals(client, auth_headers):
    resp = _upload(client, auth_headers, build_pnl())
    assert resp.status_code == 200, resp.get_data(as_text=True)
    data = resp.get_json()["data"]

    assert data["period_from"] == "2026-06-01"
    assert data["period_to"] == "2026-06-30"
    assert data["period_month"] == "2026-06-01"
    assert len(data["lines"]) == 10, "2 revenue + 4 direct + 4 indirect"

    by_name = {l["ledger_name"]: l for l in data["lines"]}
    assert by_name["RENT PAID"]["amount"] == 499885
    assert by_name["RENT PAID"]["section"] == "direct"
    assert by_name["BANK CHARGES"]["section"] == "indirect"
    assert by_name["RENT RECEIVED"]["section"] == "income"


def test_seeded_mappings_recognise_the_real_ledger_names(client, auth_headers):
    data = _upload(client, auth_headers, build_pnl()).get_json()["data"]
    unmapped = [l["ledger_name"] for l in data["lines"] if not l["is_mapped"]]
    assert unmapped == [], f"these needed manual mapping: {unmapped}"
    by_name = {l["ledger_name"]: l for l in data["lines"]}
    assert by_name["SEAWAGE REMOVAL AND CLEANING"]["category_name"] == \
        "Sewage Removal and Cleaning", "the export's misspelling is pre-mapped"
    assert by_name["RENT PAID"]["is_property_wise"] is True


def test_reconciliation_confirms_the_files_own_totals(client, auth_headers):
    data = _upload(client, auth_headers, build_pnl()).get_json()["data"]
    rec = data["reconciliation"]
    assert rec["all_match"] is True
    by_section = {c["section"]: c for c in rec["checks"]}
    assert by_section["direct"]["parsed"] == 677074.5
    assert by_section["direct"]["reported"] == 677074.5
    assert by_section["direct"]["difference"] == 0


def test_import_refuses_when_totals_disagree(client, auth_headers):
    data = _upload(client, auth_headers, build_pnl(break_totals=True)).get_json()["data"]
    assert data["reconciliation"]["all_match"] is False

    resp = client.post("/api/v1/expenses/import/post", headers=auth_headers, json={
        "file_hash": data["file_hash"],
        "mappings": {l["ledger_name"]: l["category_id"] for l in data["lines"]},
    })
    assert resp.status_code == 400
    assert "totals don't match" in resp.get_json()["message"]
    assert "Nothing was imported" in resp.get_json()["message"]

    assert client.get("/api/v1/expenses?month=2026-06-01",
                      headers=auth_headers).get_json()["meta"]["count"] == 0


def test_rejects_a_file_that_is_not_a_pnl(client, auth_headers):
    wb = Workbook()
    wb.active["A1"] = "just some spreadsheet"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = _upload(client, auth_headers, buf, "random.xlsx")
    assert resp.status_code == 400
    assert "report period" in resp.get_json()["message"]


# ---------------------------------------------------------------- posting

def _post(client, auth_headers, data, **extra):
    body = {
        "file_hash": data["file_hash"],
        "mappings": {l["ledger_name"]: l["category_id"] for l in data["lines"]},
    }
    body.update(extra)
    return client.post("/api/v1/expenses/import/post", headers=auth_headers, json=body)


def test_import_posts_expenses_and_skips_duplicated_lines(client, auth_headers):
    """RENT RECEIVED and RENT PAID both duplicate a figure the portal
    already derives structurally (RentCharge, and landlord payments/
    landlord_charges respectively) — both stay out of the expense
    ledger. OTHER INCOME has no other source in the portal and IS
    imported, so the P&L can show it."""
    data = _upload(client, auth_headers, build_pnl()).get_json()["data"]
    resp = _post(client, auth_headers, data)
    assert resp.status_code == 201, resp.get_data(as_text=True)
    batch = resp.get_json()["data"]
    assert batch["batch_number"].startswith("IMP-")
    assert batch["lines_imported"] == 8, "the two duplicated lines are not expenses"
    assert batch["reported_net_profit"] is not None

    listed = client.get("/api/v1/expenses?month=2026-06-01",
                        headers=auth_headers).get_json()
    assert listed["meta"]["count"] == 8
    names = {r["ledger_name"] for r in listed["data"]}
    assert "RENT RECEIVED" not in names, "revenue must not become an expense"
    assert "RENT PAID" not in names, "would double-count against landlord payments"
    assert "OTHER INCOME" in names, "has no other source in the portal"


def test_importing_the_same_file_twice_is_refused(client, auth_headers):
    # The same *file*, byte for byte — build_pnl() twice would produce two
    # different files, because openpyxl stamps the creation time into
    # docProps/core.xml and the hash is over the bytes.
    blob = build_pnl().getvalue()
    data = _upload(client, auth_headers, io.BytesIO(blob)).get_json()["data"]
    assert _post(client, auth_headers, data).status_code == 201

    data2 = _upload(client, auth_headers, io.BytesIO(blob)).get_json()["data"]
    assert data2["duplicate_of"] is not None
    again = _post(client, auth_headers, data2)
    assert again.status_code == 409
    assert "already imported" in again.get_json()["message"]

    forced = _post(client, auth_headers, data2, force=True)
    assert forced.status_code == 201, "an operator can override deliberately"


def test_unmapped_ledger_blocks_the_import(client, auth_headers):
    stream = build_pnl(indirect=[("A BRAND NEW LEDGER", 100)])
    data = _upload(client, auth_headers, stream).get_json()["data"]
    new_line = next(l for l in data["lines"] if l["ledger_name"] == "A BRAND NEW LEDGER")
    assert new_line["is_mapped"] is False

    resp = client.post("/api/v1/expenses/import/post", headers=auth_headers, json={
        "file_hash": data["file_hash"],
        "mappings": {l["ledger_name"]: l["category_id"]
                     for l in data["lines"] if l["category_id"]},
    })
    assert resp.status_code == 400
    assert "A BRAND NEW LEDGER" in resp.get_json()["message"]


def test_mapping_a_new_ledger_is_remembered(client, auth_headers):
    categories = client.get("/api/v1/expenses/categories",
                            headers=auth_headers).get_json()["data"]
    misc = next(c for c in categories if c["code"] == "CORPORATE_OFFICE")

    data = _upload(client, auth_headers,
                   build_pnl(indirect=[("NEW LEDGER X", 250)])).get_json()["data"]
    mappings = {l["ledger_name"]: l["category_id"] for l in data["lines"] if l["category_id"]}
    mappings["NEW LEDGER X"] = misc["id"]
    assert client.post("/api/v1/expenses/import/post", headers=auth_headers, json={
        "file_hash": data["file_hash"], "mappings": mappings}).status_code == 201

    saved = client.get("/api/v1/expenses/mappings", headers=auth_headers).get_json()["data"]
    assert any(m["ledger_name"] == "NEW LEDGER X" for m in saved)

    # Next month the same ledger arrives pre-mapped.
    july = _upload(client, auth_headers, build_pnl(
        indirect=[("NEW LEDGER X", 300)],
        period="from 01-Jul-2026 to 31-Jul-2026")).get_json()["data"]
    line = next(l for l in july["lines"] if l["ledger_name"] == "NEW LEDGER X")
    assert line["is_mapped"] is True


def test_import_can_allocate_a_line_to_a_property(client, auth_headers):
    _, prop = _property(client, auth_headers)
    data = _upload(client, auth_headers, build_pnl()).get_json()["data"]
    resp = _post(client, auth_headers, data,
                 allocations={"ELECTRICITY AND WATER FOR CAMP": prop["id"]})
    assert resp.status_code == 201

    allocated = client.get(
        f"/api/v1/expenses?month=2026-06-01&property_id={prop['id']}",
        headers=auth_headers).get_json()
    assert allocated["meta"]["count"] == 1
    assert allocated["data"][0]["amount"] == 139175


def test_voiding_a_batch_removes_its_expenses(client, auth_headers):
    data = _upload(client, auth_headers, build_pnl()).get_json()["data"]
    batch = _post(client, auth_headers, data).get_json()["data"]
    assert client.get("/api/v1/expenses?month=2026-06-01",
                      headers=auth_headers).get_json()["meta"]["count"] == 8

    resp = client.post(f"/api/v1/expenses/import/batches/{batch['id']}/void",
                       headers=auth_headers, json={"reason": "wrong month"})
    assert resp.status_code == 200
    assert client.get("/api/v1/expenses?month=2026-06-01",
                      headers=auth_headers).get_json()["meta"]["count"] == 0

    batches = client.get("/api/v1/expenses/import/batches",
                         headers=auth_headers).get_json()["data"]
    assert batches[0]["status"] == "voided", "the batch row stays for the audit trail"


# ------------------------------------------------------ landlord payments

def test_landlord_payment_voucher(client, auth_headers):
    landlord, prop = _property(client, auth_headers)
    resp = client.post("/api/v1/expenses/landlord-payments", headers=auth_headers, json={
        "landlord_id": landlord["id"], "property_id": prop["id"],
        "period_month": "2026-06-01", "amount": 30000, "mode": "cheque",
        "reference": "CHQ-4471"})
    assert resp.status_code == 201, resp.get_data(as_text=True)
    payment = resp.get_json()["data"]
    assert payment["voucher_number"].startswith("PV-")
    assert payment["amount"] == 30000
    assert payment["status"] == "posted"

    listed = client.get("/api/v1/expenses/landlord-payments?month=2026-06-01",
                        headers=auth_headers).get_json()
    assert listed["meta"]["total"] == 30000


def test_voiding_a_payment_keeps_the_voucher(client, auth_headers):
    landlord, prop = _property(client, auth_headers)
    payment = client.post("/api/v1/expenses/landlord-payments", headers=auth_headers, json={
        "landlord_id": landlord["id"], "property_id": prop["id"],
        "period_month": "2026-06-01", "amount": 30000, "mode": "cash"}).get_json()["data"]

    resp = client.post(f"/api/v1/expenses/landlord-payments/{payment['id']}/void",
                       headers=auth_headers, json={"reason": "paid twice"})
    assert resp.status_code == 200
    assert resp.get_json()["data"]["status"] == "voided"

    listed = client.get("/api/v1/expenses/landlord-payments?month=2026-06-01",
                        headers=auth_headers).get_json()
    assert listed["meta"]["count"] == 1, "still listed"
    assert listed["meta"]["total"] == 0, "but no longer counted"


def test_landlord_statement(client, auth_headers):
    landlord, prop = _property(client, auth_headers)
    for month, amount in (("2026-05-01", 30000), ("2026-06-01", 30000)):
        client.post("/api/v1/expenses/landlord-payments", headers=auth_headers, json={
            "landlord_id": landlord["id"], "property_id": prop["id"],
            "period_month": month, "amount": amount, "mode": "cheque"})

    data = client.get(f"/api/v1/expenses/landlord-statement/{landlord['id']}",
                      headers=auth_headers).get_json()["data"]
    assert data["total_paid"] == 60000
    assert len(data["payments"]) == 2


# --------------------------------------------------------- property P&L

def test_property_pnl_matches_the_workbook_block(client, auth_headers):
    """Income − rent paid − direct expenses = the block's Profit Or Loss."""
    landlord, prop = _property(client, auth_headers, "ST-10")
    tenant = client.post("/api/v1/clients", headers=auth_headers,
                         json={"name": "PNL Client"}).get_json()["data"]
    units = client.get(f"/api/v1/properties/{prop['id']}/units",
                       headers=auth_headers).get_json()["data"]
    contract = client.post("/api/v1/contracts", headers=auth_headers, json={
        "client_id": tenant["id"], "property_id": prop["id"],
        "unit_ids": [units[0]["id"]], "start_date": "2026-06-01",
        "expiry_date": "2026-12-31", "monthly_rent": 58600,
        "payment_mode": "cash"}).get_json()["data"]
    client.post("/api/v1/rent/generate", headers=auth_headers,
                json={"contract_id": contract["id"], "upto": "2026-06-01"})

    client.post("/api/v1/expenses/landlord-payments", headers=auth_headers, json={
        "landlord_id": landlord["id"], "property_id": prop["id"],
        "period_month": "2026-06-01", "amount": 30000, "mode": "cheque"})

    categories = {c["code"]: c for c in client.get(
        "/api/v1/expenses/categories", headers=auth_headers).get_json()["data"]}
    for code, amount in (("SEWAGE", 280), ("ELECTRICITY_CAMP", 11973)):
        client.post("/api/v1/expenses", headers=auth_headers, json={
            "category_id": categories[code]["id"], "property_id": prop["id"],
            "period_month": "2026-06-01", "amount": amount})

    resp = client.get("/api/v1/expenses/pnl?month=2026-06-01", headers=auth_headers)
    assert resp.status_code == 200
    pnl = resp.get_json()["data"]
    row = next(r for r in pnl["rows"] if r["property_id"] == prop["id"])

    assert row["rent_charged"] == 58600
    assert row["rent_paid"] == 30000
    assert row["expense_total"] == 12253
    assert row["profit"] == 58600 - 30000 - 12253
    assert row["expenses"]["Sewage Removal and Cleaning"] == 280


def test_property_pnl_rent_due_landlord_falls_back_to_cash_when_no_charges(client, auth_headers):
    """No landlord_charges generated for this property/month yet — the
    accrual figure falls back to the cash figure, so profit is unchanged
    from before the Phase 6 switch (the "strict no-op" guarantee)."""
    landlord, prop = _property(client, auth_headers, "ST-11")
    client.post("/api/v1/expenses/landlord-payments", headers=auth_headers, json={
        "landlord_id": landlord["id"], "property_id": prop["id"],
        "period_month": "2026-06-01", "amount": 20000, "mode": "cheque"})

    row = next(r for r in client.get(
        "/api/v1/expenses/pnl?month=2026-06-01", headers=auth_headers
    ).get_json()["data"]["rows"] if r["property_id"] == prop["id"])
    assert row["rent_paid"] == 20000
    assert row["rent_due_landlord"] == 20000, "falls back to cash with no ledger"
    assert row["profit"] == 0 - 20000 - 0


def test_property_pnl_rent_due_landlord_uses_the_dues_ledger_once_generated(client, auth_headers):
    """Once landlord_charges exist for the month, the accrual figure —
    not the cash paid — drives `profit`, even if they differ."""
    landlord, prop = _property(client, auth_headers, "ST-12")
    ag = client.post(f"/api/v1/properties/{prop['id']}/agreements", headers=auth_headers, json={
        "landlord_id": landlord["id"], "start_date": "2026-01-01", "expiry_date": "2026-12-31",
        "monthly_rent": 25000,
    }).get_json()["data"]
    client.post("/api/v1/expenses/landlord-dues/generate", headers=auth_headers,
               json={"contract_id": ag["id"], "upto": "2026-06-01"})
    # Only 20,000 actually paid — cash and accrual now genuinely differ.
    client.post("/api/v1/expenses/landlord-payments", headers=auth_headers, json={
        "landlord_id": landlord["id"], "property_id": prop["id"],
        "period_month": "2026-06-01", "amount": 20000, "mode": "cheque"})

    row = next(r for r in client.get(
        "/api/v1/expenses/pnl?month=2026-06-01", headers=auth_headers
    ).get_json()["data"]["rows"] if r["property_id"] == prop["id"])
    assert row["rent_paid"] == 20000, "cash figure untouched"
    assert row["rent_due_landlord"] == 25000, "accrual figure from the dues ledger"
    assert row["profit"] == 0 - 25000 - 0, "profit is accrual-based"
    assert row["cash_profit"] == 0 - 20000 - 0, "cash_profit stays cash-based"


def test_property_pnl_excludes_rent_paid_expense_rows(client, auth_headers):
    """A RENT_PAID Expense row (e.g. from a future import) must not
    inflate expense_total — that figure is already carried by
    rent_paid/rent_due_landlord."""
    landlord, prop = _property(client, auth_headers, "ST-13")
    categories = {c["code"]: c for c in client.get(
        "/api/v1/expenses/categories", headers=auth_headers).get_json()["data"]}
    client.post("/api/v1/expenses", headers=auth_headers, json={
        "category_id": categories["RENT_PAID"]["id"], "property_id": prop["id"],
        "period_month": "2026-06-01", "amount": 40000})

    row = next(r for r in client.get(
        "/api/v1/expenses/pnl?month=2026-06-01", headers=auth_headers
    ).get_json()["data"]["rows"] if r["property_id"] == prop["id"])
    assert row["expense_total"] == 0, "RENT_PAID must not land in direct expenses"
    assert "Rent Paid" not in row["expenses"]


def test_property_pnl_excludes_opening_balance_from_rent_charged(client, auth_headers):
    """A carried-forward balance is booked once, dated the month before
    the tenancy starts — it must not be double-counted as this month's
    rent income wherever its period_month happens to land."""
    landlord, prop = _property(client, auth_headers, "ST-14")
    tenant = client.post("/api/v1/clients", headers=auth_headers,
                         json={"name": "Opening Balance Client"}).get_json()["data"]
    units = client.get(f"/api/v1/properties/{prop['id']}/units",
                       headers=auth_headers).get_json()["data"]
    contract = client.post("/api/v1/contracts", headers=auth_headers, json={
        "client_id": tenant["id"], "property_id": prop["id"],
        "unit_ids": [units[0]["id"]], "start_date": "2026-07-01",
        "expiry_date": "2026-12-31", "monthly_rent": 5000,
        "opening_balance": 15000, "payment_mode": "cash"}).get_json()["data"]
    client.post("/api/v1/rent/generate", headers=auth_headers,
               json={"contract_id": contract["id"], "upto": "2026-07-01"})

    charges = {c["period_month"]: c for c in client.get(
        f"/api/v1/rent/charges?contract_id={contract['id']}",
        headers=auth_headers).get_json()["data"]}
    opening_month = min(charges)  # dated the month before start_date, per rent.py
    assert charges[opening_month]["is_opening_balance"] is True

    row = next(r for r in client.get(
        f"/api/v1/expenses/pnl?month={opening_month}", headers=auth_headers
    ).get_json()["data"]["rows"] if r["property_id"] == prop["id"])
    assert row["rent_charged"] == 0, "the opening-balance charge itself is excluded"


def test_property_pnl_separates_company_overhead(client, auth_headers):
    _, prop = _property(client, auth_headers)
    categories = {c["code"]: c for c in client.get(
        "/api/v1/expenses/categories", headers=auth_headers).get_json()["data"]}

    client.post("/api/v1/expenses", headers=auth_headers, json={
        "category_id": categories["SEWAGE"]["id"], "property_id": prop["id"],
        "period_month": "2026-06-01", "amount": 500})
    client.post("/api/v1/expenses", headers=auth_headers, json={
        "category_id": categories["SALARY"]["id"],
        "period_month": "2026-06-01", "amount": 65900})

    pnl = client.get("/api/v1/expenses/pnl?month=2026-06-01",
                     headers=auth_headers).get_json()["data"]
    assert pnl["totals"]["expense_total"] == 500, "only property costs hit the property"
    assert pnl["totals"]["company_overhead"] == 65900
    assert pnl["totals"]["net_profit"] == -500 - 65900


def test_unallocated_direct_cost_is_flagged_not_hidden(client, auth_headers):
    _property(client, auth_headers)
    categories = {c["code"]: c for c in client.get(
        "/api/v1/expenses/categories", headers=auth_headers).get_json()["data"]}
    client.post("/api/v1/expenses", headers=auth_headers, json={
        "category_id": categories["ELECTRICITY_CAMP"]["id"],
        "period_month": "2026-06-01", "amount": 139175})

    pnl = client.get("/api/v1/expenses/pnl?month=2026-06-01",
                     headers=auth_headers).get_json()["data"]
    assert pnl["totals"]["unallocated_direct"] == 139175
    assert pnl["totals"]["net_profit"] == -139175


def test_allocating_an_imported_cost_moves_it_into_the_property(client, auth_headers):
    _, prop = _property(client, auth_headers)
    data = _upload(client, auth_headers, build_pnl()).get_json()["data"]
    _post(client, auth_headers, data)

    unallocated = client.get("/api/v1/expenses?month=2026-06-01&unallocated=1",
                             headers=auth_headers).get_json()["data"]
    electricity = next(r for r in unallocated
                       if r["ledger_name"] == "ELECTRICITY AND WATER FOR CAMP")

    resp = client.post(f"/api/v1/expenses/{electricity['id']}/allocate",
                       headers=auth_headers, json={"property_id": prop["id"]})
    assert resp.status_code == 200

    pnl = client.get("/api/v1/expenses/pnl?month=2026-06-01",
                     headers=auth_headers).get_json()["data"]
    row = next(r for r in pnl["rows"] if r["property_id"] == prop["id"])
    assert row["expense_total"] == 139175


def test_expense_endpoints_require_auth(client):
    assert client.get("/api/v1/expenses").status_code == 401
    assert client.get("/api/v1/expenses/pnl").status_code == 401
    assert client.post("/api/v1/expenses/import/preview").status_code == 401
