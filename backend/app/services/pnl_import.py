"""Import the accounting software's Profit & Loss export.

The export is a printed report, not a data feed: a title carrying the
period, indented ledger names in column A, line amounts in column B and
section totals in column C. We parse it, and then — the part that makes
the import trustworthy — **check our sums against the totals the file
prints for itself**. If they disagree we refuse to post rather than
quietly load half a month.

Ledger names are mapped to categories once and the mapping is reused
every month, so the second import is a two-click job.
"""
from __future__ import annotations

import hashlib
import re
from datetime import date, datetime

from openpyxl import load_workbook

from ..extensions import db
from ..models import Expense, ExpenseCategory, ExpenseImportBatch, LedgerMapping
from . import audit


class ImportError_(ValueError):
    """Caller-visible failure. Named with a trailing underscore so it
    doesn't shadow the builtin ImportError."""


# Section headings as the report prints them, mapped to the kind of line
# that follows. Compared case-insensitively after stripping.
SECTION_HEADERS = {
    "revenues": "income",
    "purchase & direct expenses": "direct",
    "indirect expenses": "indirect",
}

# Rows that are totals rather than ledger lines. Used both to skip them
# and to reconcile against.
TOTAL_LABELS = {
    "total revenues": "revenue",
    "cost of sales": "direct",
    "(gross profit)": "gross_profit",
    "total indirect expenses": "indirect",
    "(net profit)": "net_profit",
}

PERIOD_RE = re.compile(
    r"from\s+(\d{1,2}-\w{3}-\d{4})\s+to\s+(\d{1,2}-\w{3}-\d{4})", re.IGNORECASE)


def _parse_report_date(value: str) -> date:
    return datetime.strptime(value.strip(), "%d-%b-%Y").date()


def _number(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return None


def parse_workbook(stream, *, original_name: str = "") -> dict:
    """Read the export into ledger lines plus the totals it declares.

    Returns ``{period_from, period_to, period_month, lines, reported,
    file_hash}`` where each line is
    ``{section, ledger_name, amount, row}``.
    """
    raw = stream.read() if hasattr(stream, "read") else stream
    if not raw:
        raise ImportError_("The file is empty")
    file_hash = hashlib.sha256(raw).hexdigest()

    import io as _io
    try:
        wb = load_workbook(_io.BytesIO(raw), data_only=True)
    except Exception as exc:  # openpyxl raises a family of errors
        raise ImportError_(f"Could not read the workbook: {exc}") from exc

    ws = wb[wb.sheetnames[0]]

    period_from = period_to = None
    section = None
    lines: list[dict] = []
    reported: dict[str, float] = {}

    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if not isinstance(label, str):
            continue
        text = label.strip()
        if not text:
            continue

        if period_from is None:
            found = PERIOD_RE.search(text)
            if found:
                try:
                    period_from = _parse_report_date(found.group(1))
                    period_to = _parse_report_date(found.group(2))
                except ValueError:
                    raise ImportError_(
                        f"Could not read the report period from: {text!r}")
                continue

        key = text.lower()

        if key in TOTAL_LABELS:
            value = _number(ws.cell(row, 3).value) or _number(ws.cell(row, 2).value)
            if value is not None:
                reported[TOTAL_LABELS[key]] = value
            continue

        if key in SECTION_HEADERS:
            section = SECTION_HEADERS[key]
            continue

        amount = _number(ws.cell(row, 2).value)
        if amount is None:
            continue
        if section is None:
            # A figure before any section heading — the layout isn't what
            # we expect, so say so rather than guessing.
            raise ImportError_(
                f"Row {row}: found an amount for {text!r} before any section heading")
        lines.append({
            "section": section,
            "ledger_name": text,
            "amount": amount,
            "row": row,
        })

    if period_from is None or period_to is None:
        raise ImportError_(
            "Could not find the report period. Expected a title like "
            "'... from 01-Jun-2026 to 30-Jun-2026'.")
    if not lines:
        raise ImportError_("No ledger lines found in the file")

    return {
        "period_from": period_from,
        "period_to": period_to,
        "period_month": period_from.replace(day=1),
        "lines": lines,
        "reported": reported,
        "file_hash": file_hash,
        "original_name": original_name,
    }


def reconcile(parsed: dict) -> dict:
    """Compare what we parsed with what the file says its totals are."""
    sums = {"income": 0.0, "direct": 0.0, "indirect": 0.0}
    for line in parsed["lines"]:
        sums[line["section"]] = round(sums[line["section"]] + line["amount"], 2)

    reported = parsed["reported"]
    checks = []
    for section, reported_key, label in (
        ("income", "revenue", "Total Revenues"),
        ("direct", "direct", "Cost of Sales"),
        ("indirect", "indirect", "Total Indirect Expenses"),
    ):
        expected = reported.get(reported_key)
        got = sums[section]
        checks.append({
            "section": section,
            "label": label,
            "parsed": got,
            "reported": expected,
            "matches": expected is None or abs(expected - got) < 0.01,
            "difference": None if expected is None else round(got - expected, 2),
        })

    net = None
    if all(k in reported for k in ("revenue", "direct", "indirect")):
        net = round(reported["revenue"] - reported["direct"] - reported["indirect"], 2)

    return {
        "sums": sums,
        "checks": checks,
        "all_match": all(c["matches"] for c in checks),
        "computed_net_profit": net,
        "reported_net_profit": reported.get("net_profit"),
    }


def suggest_mappings(parsed: dict) -> list[dict]:
    """Attach the saved category for each ledger name, where known."""
    names = [l["ledger_name"] for l in parsed["lines"]]
    saved = {
        m.ledger_name.lower(): m
        for m in LedgerMapping.query.filter(
            db.func.lower(LedgerMapping.ledger_name).in_([n.lower() for n in names])
        ).all()
    }
    out = []
    for line in parsed["lines"]:
        mapping = saved.get(line["ledger_name"].lower())
        category = mapping.category if mapping else None
        out.append({
            **line,
            "category_id": category.id if category else None,
            "category_name": category.name if category else None,
            "category_kind": category.kind if category else None,
            "is_property_wise": category.is_property_wise if category else False,
            "is_mapped": category is not None,
        })
    return out


def _next_batch_number() -> str:
    month_key = datetime.utcnow().strftime("%Y%m")
    like = f"IMP-{month_key}-%"
    last = (
        db.session.query(ExpenseImportBatch.batch_number)
        .filter(ExpenseImportBatch.batch_number.like(like))
        .order_by(ExpenseImportBatch.batch_number.desc())
        .limit(1)
        .scalar()
    )
    seq = 1
    if last:
        try:
            seq = int(last.rsplit("-", 1)[1]) + 1
        except (IndexError, ValueError):
            seq = 1
    return f"IMP-{month_key}-{seq:04d}"


def find_duplicate(parsed: dict) -> ExpenseImportBatch | None:
    return (
        ExpenseImportBatch.query
        .filter_by(file_hash=parsed["file_hash"], status="posted")
        .first()
    )


def post_import(parsed: dict, *, mappings: dict[str, int], actor,
                allocations: dict[str, int] | None = None,
                force: bool = False, remarks: str | None = None) -> ExpenseImportBatch:
    """Store the parsed month.

    `mappings` is ``{ledger_name: category_id}`` — saved for reuse.
    `allocations` optionally pins a ledger line to a property.
    """
    check = reconcile(parsed)
    if not check["all_match"] and not force:
        bad = [c for c in check["checks"] if not c["matches"]]
        detail = "; ".join(
            f"{c['label']}: file says {c['reported']}, lines add to {c['parsed']}"
            for c in bad)
        raise ImportError_(
            f"The file's own totals don't match its lines — {detail}. "
            "Nothing was imported.")

    duplicate = find_duplicate(parsed)
    if duplicate is not None and not force:
        raise ImportError_(
            f"This exact file was already imported as {duplicate.batch_number} "
            f"on {duplicate.created_at:%Y-%m-%d}. Nothing was imported.")

    unmapped = [l["ledger_name"] for l in parsed["lines"]
                if not mappings.get(l["ledger_name"])]
    if unmapped:
        raise ImportError_(
            "These ledgers have no category yet: " + ", ".join(sorted(unmapped)))

    reported = parsed["reported"]
    batch = ExpenseImportBatch(
        batch_number=_next_batch_number(),
        original_name=parsed.get("original_name") or "import.xlsx",
        file_hash=parsed["file_hash"],
        period_from=parsed["period_from"],
        period_to=parsed["period_to"],
        period_month=parsed["period_month"],
        reported_revenue=reported.get("revenue"),
        reported_direct=reported.get("direct"),
        reported_indirect=reported.get("indirect"),
        reported_net_profit=reported.get("net_profit"),
        status="posted",
        remarks=remarks,
        created_by=actor.id,
        updated_by=actor.id,
    )
    db.session.add(batch)
    db.session.flush()

    # These two ledger lines duplicate figures the portal already derives
    # structurally — RENT_RECEIVED from client RentCharge/receipts,
    # RENT_PAID from landlord payments/landlord_charges (Phase 4).
    # Importing either as an Expense row would double-count it in
    # property_pnl()/company-pnl. Every other income line (Other Income,
    # etc.) has no other source in the portal and is imported normally.
    DUPLICATE_CODES = {"RENT_RECEIVED", "RENT_PAID"}

    allocations = allocations or {}
    stored = 0
    skipped: list[str] = []
    for line in parsed["lines"]:
        category_id = mappings[line["ledger_name"]]
        category = ExpenseCategory.query.get(category_id)
        if category is None:
            raise ImportError_(f"Category {category_id} not found")

        # Remember the mapping for next month.
        existing = LedgerMapping.query.filter(
            db.func.lower(LedgerMapping.ledger_name) == line["ledger_name"].lower()
        ).first()
        if existing is None:
            db.session.add(LedgerMapping(
                ledger_name=line["ledger_name"], category_id=category_id,
                created_by=actor.id, updated_by=actor.id))
        elif existing.category_id != category_id:
            existing.category_id = category_id
            existing.updated_by = actor.id

        if category.code in DUPLICATE_CODES:
            skipped.append(f"{line['ledger_name']} ({category.code}) — "
                           f"already tracked elsewhere in the portal, not imported")
            continue

        db.session.add(Expense(
            category_id=category_id,
            property_id=allocations.get(line["ledger_name"]),
            period_month=parsed["period_month"],
            amount=line["amount"],
            source="import",
            batch_id=batch.id,
            ledger_name=line["ledger_name"],
            created_by=actor.id,
            updated_by=actor.id,
        ))
        stored += 1

    batch.lines_imported = stored
    if skipped:
        note = "Skipped (double-count guard): " + "; ".join(skipped)
        batch.remarks = f"{batch.remarks}\n{note}" if batch.remarks else note
    db.session.flush()

    audit.record(user=actor, action="import", module="expense",
                 entity_type="expense_import_batch", entity_id=batch.id,
                 new_value={
                     "batch": batch.batch_number,
                     "period": parsed["period_month"].isoformat(),
                     "lines": stored,
                     "skipped": skipped,
                     "reconciled": check["all_match"],
                 },
                 remarks=f"imported {parsed.get('original_name')}")
    return batch


def void_batch(batch: ExpenseImportBatch, *, reason: str, actor) -> ExpenseImportBatch:
    """Reverse a whole import. The batch row stays for the audit trail."""
    if batch.status == "voided":
        raise ImportError_(f"Batch {batch.batch_number} is already voided")
    if not (reason or "").strip():
        raise ImportError_("reason is required to void an import")

    for expense in list(batch.expenses or []):
        db.session.delete(expense)
    batch.status = "voided"
    batch.remarks = ((batch.remarks + "\n") if batch.remarks else "") + f"Voided: {reason}"
    batch.updated_by = actor.id
    db.session.flush()

    audit.record(user=actor, action="void", module="expense",
                 entity_type="expense_import_batch", entity_id=batch.id,
                 old_value={"status": "posted"},
                 new_value={"status": "voided", "reason": reason})
    return batch
