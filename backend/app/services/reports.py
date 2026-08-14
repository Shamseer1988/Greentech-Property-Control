"""Report registry and shared filter helpers.

Each report is a function returning ``{"columns": [...], "rows": [...], "meta": {...}}``.
The route layer turns the report into JSON, an Excel workbook or a PDF.

Reports that restate a figure already computed elsewhere call that
service rather than re-querying — the ageing report *is* `ageing()`, the
property P&L *is* `property_pnl()`. A report and the page it mirrors can
therefore never disagree.
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import func, case

from ..extensions import db
from ..models import (
    AuditLog, Cheque, Client, ClientContract, ContractUnit, Expense,
    ExpenseCategory, Floor, Landlord, Property, PropertyAgreement,
    RentCharge, Unit,
)
from .reminders import agreement_bucket
from .documents import expiring_documents
from .rent import month_start, month_end, add_months
from .ageing import ageing as ageing_service
from .expenses import property_pnl as property_pnl_service
from .risk import client_risk_report


Column = dict  # {"key": str, "label": str, "width": int | None}

REPORT_REGISTRY: dict[str, dict] = {}


def report(slug: str, title: str, category: str, description: str = ""):
    """Register a report builder."""

    def decorator(fn: Callable):
        REPORT_REGISTRY[slug] = {
            "slug": slug, "title": title, "category": category,
            "description": description, "builder": fn,
        }
        return fn

    return decorator


def list_reports() -> list[dict]:
    return [
        {k: v for k, v in r.items() if k != "builder"}
        for r in sorted(REPORT_REGISTRY.values(), key=lambda r: (r["category"], r["title"]))
    ]


def build_report(slug: str, filters: dict) -> dict:
    info = REPORT_REGISTRY.get(slug)
    if info is None:
        raise KeyError(f"Unknown report: {slug}")
    return info["builder"](filters)


# ---------- Shared filter parsing ----------

def parse_date(value, fallback: date | None = None) -> date | None:
    if value is None or value == "":
        return fallback
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    return datetime.fromisoformat(str(value)).date()


# ---------- Excel writer ----------

def to_workbook(title: str, columns: list[Column], rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1f3a8a")
    header_align = Alignment(horizontal="left", vertical="center")
    ws.append([c["label"] for c in columns])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for r in rows:
        ws.append([_excel_value(r.get(c["key"])) for c in columns])

    for idx, c in enumerate(columns, start=1):
        width = c.get("width") or max(len(c["label"]) + 2, 12)
        ws.column_dimensions[get_column_letter(idx)].width = min(width, 50)

    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _excel_value(v):
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    if isinstance(v, bool):
        return "yes" if v else "no"
    return v


# ---------- PDF writer ----------

def to_pdf(title: str, columns: list[Column], rows: list[dict],
           *, subtitle: str | None = None) -> bytes:
    """Render a report as a landscape A4 PDF — the version that gets
    printed, signed and filed.

    Column widths come from the same `width` hint the Excel writer uses,
    scaled to the printable area, so both exports lay out alike.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer)

    buf = io.BytesIO()
    page = landscape(A4)
    doc = SimpleDocTemplate(
        buf, pagesize=page,
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=title, author="GreenTech Trading & Contracting",
    )

    styles = getSampleStyleSheet()
    cell = ParagraphStyle("cell", parent=styles["BodyText"], fontSize=7,
                          leading=9, spaceAfter=0)
    head = ParagraphStyle("head", parent=cell, textColor=colors.white,
                          fontName="Helvetica-Bold")

    story = [Paragraph(title, styles["Heading2"])]
    if subtitle:
        story.append(Paragraph(subtitle, styles["Normal"]))
    story.append(Spacer(1, 6))

    if not rows:
        story.append(Paragraph("No data for these filters.", styles["Normal"]))
        doc.build(story)
        return buf.getvalue()

    available = page[0] - doc.leftMargin - doc.rightMargin
    hints = [float(c.get("width") or 14) for c in columns]
    total_hint = sum(hints) or 1
    widths = [available * h / total_hint for h in hints]

    data = [[Paragraph(str(c["label"]), head) for c in columns]]
    for r in rows:
        data.append([Paragraph(_pdf_value(r.get(c["key"])), cell) for c in columns])

    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3a8a")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#c9ced9")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f2f4f8")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(table)
    doc.build(story)
    return buf.getvalue()


def _pdf_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "yes" if v else "no"
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    if isinstance(v, float):
        return f"{v:,.2f}"
    if isinstance(v, int):
        return f"{v:,}"
    # Paragraph parses its input as mini-HTML, so anything the data
    # happens to contain must be escaped or the build fails.
    return (str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


# ============================================================
# Reports
# ============================================================

# 1. Property Occupancy (unit level)
@report(
    "property-occupancy",
    "Property Occupancy Report",
    "Occupancy",
    "Per-property unit totals, occupancy % and a breakdown by unit status.",
)
def _r_property_occupancy(filters: dict) -> dict:
    q = (
        db.session.query(
            Property.id, Property.code, Property.name, Property.property_type,
            Property.city, Property.status,
            func.count(Unit.id).label("total"),
            func.sum(case((Unit.occupancy_status == "occupied", 1), else_=0)).label("occupied"),
            func.sum(case((Unit.occupancy_status == "empty", 1), else_=0)).label("empty"),
            func.sum(case((Unit.occupancy_status == "maintenance", 1), else_=0)).label("maintenance"),
            func.sum(case((Unit.occupancy_status == "blocked", 1), else_=0)).label("blocked"),
        )
        .outerjoin(Unit, Unit.property_id == Property.id)
        .group_by(Property.id)
    )
    if filters.get("status"):
        q = q.filter(Property.status == filters["status"])
    if filters.get("city"):
        q = q.filter(db.func.lower(Property.city) == filters["city"].lower())

    rows = []
    for r in q.all():
        total = int(r.total or 0)
        occ = int(r.occupied or 0)
        rows.append({
            "code": r.code, "name": r.name, "property_type": r.property_type,
            "city": r.city, "status": r.status,
            "total": total,
            "occupied": occ,
            "empty": int(r.empty or 0),
            "maintenance": int(r.maintenance or 0),
            "blocked": int(r.blocked or 0),
            "occupancy_percent": round(occ * 100 / total, 1) if total else 0.0,
        })
    rows.sort(key=lambda r: r["occupancy_percent"], reverse=True)
    return {
        "columns": [
            {"key": "code", "label": "Code", "width": 14},
            {"key": "name", "label": "Property", "width": 28},
            {"key": "property_type", "label": "Type", "width": 18},
            {"key": "city", "label": "City", "width": 16},
            {"key": "status", "label": "Status", "width": 12},
            {"key": "total", "label": "Total Units"},
            {"key": "occupied", "label": "Occupied"},
            {"key": "empty", "label": "Empty"},
            {"key": "maintenance", "label": "Maintenance"},
            {"key": "blocked", "label": "Blocked"},
            {"key": "occupancy_percent", "label": "Occupancy %"},
        ],
        "rows": rows,
        "meta": {"count": len(rows)},
    }


# 2. Empty Units
@report(
    "empty-units",
    "Empty Units Report",
    "Occupancy",
    "Every vacant unit with its property, floor and unit number.",
)
def _r_empty_units(filters: dict) -> dict:
    q = (
        db.session.query(Unit, Floor, Property)
        .join(Floor, Unit.floor_id == Floor.id)
        .join(Property, Unit.property_id == Property.id)
        .filter(Unit.occupancy_status == "empty")
    )
    if filters.get("property_id"):
        q = q.filter(Unit.property_id == int(filters["property_id"]))

    rows = []
    for unit, floor, prop in q.order_by(Property.code, Floor.floor_number, Unit.unit_number).all():
        rows.append({
            "property_code": prop.code,
            "property_name": prop.name,
            "floor": floor.floor_number,
            "unit_number": unit.unit_number,
            "unit_type": unit.unit_type,
            "monthly_rent": float(unit.monthly_rent) if unit.monthly_rent is not None else None,
        })
    return {
        "columns": [
            {"key": "property_code", "label": "Code", "width": 14},
            {"key": "property_name", "label": "Property", "width": 28},
            {"key": "floor", "label": "Floor", "width": 10},
            {"key": "unit_number", "label": "Unit No.", "width": 12},
            {"key": "unit_type", "label": "Type", "width": 14},
            {"key": "monthly_rent", "label": "Monthly Rent", "width": 14},
        ],
        "rows": rows,
        "meta": {"count": len(rows)},
    }


# 2b. Empty Units Trend
@report(
    "empty-units-trend",
    "Empty Units Trend",
    "Occupancy",
    "How many units sat empty each month, property-wise and total — "
    "derived from contract unit history, not a live snapshot.",
)
def _r_empty_units_trend(filters: dict) -> dict:
    end = month_start(parse_date(filters.get("end"), date.today()))
    start = parse_date(filters.get("start"), None)
    start = month_start(start) if start else add_months(end, -5)
    if start > end:
        start, end = end, start

    prop_q = Unit.query.join(Property, Unit.property_id == Property.id)
    if filters.get("property_id"):
        prop_q = prop_q.filter(Unit.property_id == int(filters["property_id"]))

    totals: dict[int, dict] = {}
    for unit in prop_q.all():
        entry = totals.setdefault(unit.property_id, {"total": 0})
        entry["total"] += 1

    properties = {
        p.id: p for p in Property.query.filter(Property.id.in_(totals.keys())).all()
    } if totals else {}

    months = []
    cursor = start
    while cursor <= end:
        months.append(cursor)
        cursor = add_months(cursor, 1)

    rows = []
    for month in months:
        m_end = month_end(month)
        occ_q = (
            db.session.query(ContractUnit.unit_id, Unit.property_id)
            .join(Unit, ContractUnit.unit_id == Unit.id)
            .filter(ContractUnit.from_date <= m_end)
            .filter((ContractUnit.to_date.is_(None)) | (ContractUnit.to_date >= month))
        )
        if filters.get("property_id"):
            occ_q = occ_q.filter(Unit.property_id == int(filters["property_id"]))

        occupied_by_property: dict[int, set] = {}
        for unit_id, property_id in occ_q.all():
            occupied_by_property.setdefault(property_id, set()).add(unit_id)

        for property_id, entry in totals.items():
            occupied = len(occupied_by_property.get(property_id, ()))
            total = entry["total"]
            empty = max(total - occupied, 0)
            prop = properties.get(property_id)
            rows.append({
                "month": month.isoformat(),
                "property_id": property_id,
                "property_code": prop.code if prop else None,
                "property_name": prop.name if prop else None,
                "total_units": total,
                "occupied_units": occupied,
                "empty_units": empty,
                "empty_percent": round(empty * 100 / total, 1) if total else 0.0,
            })

    rows.sort(key=lambda r: (r["month"], r["property_code"] or ""))
    return {
        "columns": [
            {"key": "month", "label": "Month", "width": 12},
            {"key": "property_code", "label": "Code", "width": 14},
            {"key": "property_name", "label": "Property", "width": 26},
            {"key": "total_units", "label": "Total Units"},
            {"key": "occupied_units", "label": "Occupied"},
            {"key": "empty_units", "label": "Empty"},
            {"key": "empty_percent", "label": "Empty %"},
        ],
        "rows": rows,
        "meta": {"count": len(rows), "start": start.isoformat(), "end": end.isoformat()},
    }


# 3. Landlord Agreement Expiry
@report(
    "agreement-expiry",
    "Landlord Agreement Expiry Report",
    "Contracts",
    "Active landlord agreements with expiry dates and reminder buckets.",
)
def _r_agreement_expiry(filters: dict) -> dict:
    q = (
        db.session.query(PropertyAgreement, Property, Landlord)
        .join(Property, PropertyAgreement.property_id == Property.id)
        .join(Landlord, PropertyAgreement.landlord_id == Landlord.id)
        .filter(PropertyAgreement.is_active.is_(True))
    )
    if filters.get("property_id"):
        q = q.filter(PropertyAgreement.property_id == int(filters["property_id"]))

    today = date.today()
    rows = []
    for ag, prop, ll in q.order_by(PropertyAgreement.expiry_date.asc()).all():
        rows.append({
            "property_code": prop.code,
            "property_name": prop.name,
            "landlord": ll.name,
            "agreement_number": ag.agreement_number,
            "start_date": ag.start_date,
            "expiry_date": ag.expiry_date,
            "monthly_rent": float(ag.monthly_rent) if ag.monthly_rent is not None else None,
            "days_left": (ag.expiry_date - today).days,
            "bucket": agreement_bucket(ag.expiry_date, today),
            "renewal_status": ag.renewal_status,
        })
    if filters.get("bucket"):
        rows = [r for r in rows if r["bucket"] == filters["bucket"]]
    return {
        "columns": [
            {"key": "property_code", "label": "Code", "width": 14},
            {"key": "property_name", "label": "Property", "width": 28},
            {"key": "landlord", "label": "Landlord", "width": 28},
            {"key": "agreement_number", "label": "Agreement #", "width": 16},
            {"key": "start_date", "label": "Start", "width": 12},
            {"key": "expiry_date", "label": "Expiry", "width": 12},
            {"key": "monthly_rent", "label": "Monthly Rent", "width": 14},
            {"key": "days_left", "label": "Days Left"},
            {"key": "bucket", "label": "Bucket", "width": 10},
            {"key": "renewal_status", "label": "Renewal", "width": 12},
        ],
        "rows": rows,
        "meta": {"count": len(rows)},
    }


# 4. Document Expiry
@report(
    "document-expiry",
    "Document Expiry Report",
    "Contracts",
    "Company and government documents (CR / QID) already expired or expiring soon.",
)
def _r_document_expiry(filters: dict) -> dict:
    within = int(filters.get("within_days") or 90)
    rows = expiring_documents(within_days=within)
    if filters.get("entity_type"):
        rows = [r for r in rows if r["entity_type"] == filters["entity_type"]]
    return {
        "columns": [
            {"key": "entity_type", "label": "Type", "width": 12},
            {"key": "entity_name", "label": "Belongs to", "width": 28},
            {"key": "category", "label": "Document", "width": 14},
            {"key": "doc_number", "label": "Doc No.", "width": 18},
            {"key": "original_name", "label": "File", "width": 26},
            {"key": "expiry_date", "label": "Expiry", "width": 12},
            {"key": "days_left", "label": "Days Left"},
            {"key": "bucket", "label": "Bucket", "width": 10},
        ],
        "rows": rows,
        "meta": {"count": len(rows), "within_days": within},
    }


# ============================================================
# The reports that replace the workbook's own sheets
# ============================================================

# 5. Rent Collection
@report(
    "rent-collection",
    "Rent Collection Report",
    "Money",
    "Charged, collected and outstanding per client contract for a month.",
)
def _r_rent_collection(filters: dict) -> dict:
    period = month_start(parse_date(filters.get("month"), date.today()))
    q = (
        db.session.query(RentCharge, ClientContract, Client, Property)
        .join(ClientContract, RentCharge.contract_id == ClientContract.id)
        .join(Client, RentCharge.client_id == Client.id)
        .join(Property, RentCharge.property_id == Property.id)
        .filter(RentCharge.period_month == period)
    )
    if filters.get("property_id"):
        q = q.filter(RentCharge.property_id == int(filters["property_id"]))
    if filters.get("status"):
        q = q.filter(RentCharge.status == filters["status"])

    rows = []
    for charge, contract, client, prop in q.order_by(Property.code, Client.name).all():
        amount = float(charge.amount or 0)
        allocated = float(charge.allocated or 0)
        rows.append({
            "property_code": prop.code,
            "property_name": prop.name,
            "client_name": client.name,
            "contract_number": contract.contract_number,
            "payment_mode": contract.payment_mode,
            "charged": round(amount, 2),
            "collected": round(allocated, 2),
            "outstanding": round(amount - allocated, 2),
            "status": charge.status,
            "period_month": charge.period_month,
        })
    charged = round(sum(r["charged"] for r in rows), 2)
    collected = round(sum(r["collected"] for r in rows), 2)
    return {
        "columns": [
            {"key": "property_code", "label": "Code", "width": 14},
            {"key": "property_name", "label": "Property", "width": 26},
            {"key": "client_name", "label": "Client", "width": 28},
            {"key": "contract_number", "label": "Contract", "width": 18},
            {"key": "payment_mode", "label": "Mode", "width": 10},
            {"key": "charged", "label": "Charged", "width": 14},
            {"key": "collected", "label": "Collected", "width": 14},
            {"key": "outstanding", "label": "Outstanding", "width": 14},
            {"key": "status", "label": "Status", "width": 12},
            {"key": "period_month", "label": "Month", "width": 12},
        ],
        "rows": rows,
        "meta": {
            "count": len(rows),
            "month": period.isoformat(),
            "charged": charged,
            "collected": collected,
            "outstanding": round(charged - collected, 2),
            "collection_percent": round(collected * 100 / charged, 1) if charged else 0.0,
        },
    }


# 6. Ageing
@report(
    "ageing",
    "Ageing Report",
    "Money",
    "Outstanding per client per month, oldest first — the Ageing sheet.",
)
def _r_ageing(filters: dict) -> dict:
    upto = month_start(parse_date(filters.get("upto"), date.today()))
    months = int(filters.get("months") or 12)
    property_id = int(filters["property_id"]) if filters.get("property_id") else None
    data = ageing_service(upto=upto, months=months, property_id=property_id)

    # One column per month in the window, so the sheet reads across like
    # the workbook did, with anything older rolled into a single column.
    columns = [
        {"key": "client_code", "label": "Code", "width": 14},
        {"key": "client_name", "label": "Client", "width": 30},
        {"key": "older", "label": "Older", "width": 14},
    ]
    columns += [{"key": m, "label": m[:7], "width": 13} for m in data["months"]]
    columns.append({"key": "total", "label": "Total Due", "width": 15})

    rows = []
    for entry in data["rows"]:
        row = {
            "client_code": entry["client_code"],
            "client_name": entry["client_name"],
            "older": entry["older"] or None,
            "total": entry["total"],
        }
        for month in data["months"]:
            row[month] = entry["months"].get(month) or None
        rows.append(row)
    return {
        "columns": columns,
        "rows": rows,
        "meta": {
            "count": len(rows),
            "upto": data["upto"],
            "grand_total": data["grand_total"],
        },
    }


# 7. Client Contract Expiry (the Expiry Control sheet)
@report(
    "contract-expiry",
    "Contract Expiry Control",
    "Contracts",
    "Active client contracts by expiry date, with the renewal window.",
)
def _r_contract_expiry(filters: dict) -> dict:
    today = date.today()
    within = int(filters.get("within_days") or 90)
    q = (
        db.session.query(ClientContract, Client, Property)
        .join(Client, ClientContract.client_id == Client.id)
        .join(Property, ClientContract.property_id == Property.id)
        .filter(ClientContract.status == "active")
    )
    if not filters.get("all"):
        q = q.filter(ClientContract.expiry_date <= today + timedelta(days=within))
    if filters.get("property_id"):
        q = q.filter(ClientContract.property_id == int(filters["property_id"]))

    rows = []
    for contract, client, prop in q.order_by(ClientContract.expiry_date).all():
        days = (contract.expiry_date - today).days
        rows.append({
            "property_code": prop.code,
            "property_name": prop.name,
            "client_name": client.name,
            "contract_number": contract.contract_number,
            "start_date": contract.start_date,
            "expiry_date": contract.expiry_date,
            "days_left": days,
            "bucket": ("expired" if days < 0 else "30 days" if days <= 30
                       else "60 days" if days <= 60 else "90 days"),
            "monthly_rent": float(contract.monthly_rent),
            "payment_mode": contract.payment_mode,
        })
    return {
        "columns": [
            {"key": "property_code", "label": "Code", "width": 14},
            {"key": "property_name", "label": "Property", "width": 26},
            {"key": "client_name", "label": "Client", "width": 28},
            {"key": "contract_number", "label": "Contract", "width": 18},
            {"key": "start_date", "label": "Start", "width": 12},
            {"key": "expiry_date", "label": "Expiry", "width": 12},
            {"key": "days_left", "label": "Days Left"},
            {"key": "bucket", "label": "Bucket", "width": 10},
            {"key": "monthly_rent", "label": "Monthly Rent", "width": 14},
            {"key": "payment_mode", "label": "Mode", "width": 10},
        ],
        "rows": rows,
        "meta": {"count": len(rows), "within_days": within},
    }


# 8. PDC Register
@report(
    "pdc-register",
    "PDC Register",
    "Money",
    "Post-dated cheques on hand with their status and banking dates.",
)
def _r_pdc_register(filters: dict) -> dict:
    q = (
        db.session.query(Cheque, ClientContract, Client, Property)
        .join(ClientContract, Cheque.contract_id == ClientContract.id)
        .join(Client, ClientContract.client_id == Client.id)
        .join(Property, ClientContract.property_id == Property.id)
    )
    if filters.get("status"):
        q = q.filter(Cheque.status == filters["status"])
    if filters.get("property_id"):
        q = q.filter(ClientContract.property_id == int(filters["property_id"]))
    if filters.get("from"):
        q = q.filter(Cheque.cheque_date >= parse_date(filters["from"]))
    if filters.get("to"):
        q = q.filter(Cheque.cheque_date <= parse_date(filters["to"]))

    rows = []
    for cheque, contract, client, prop in q.order_by(Cheque.cheque_date).all():
        rows.append({
            "cheque_date": cheque.cheque_date,
            "cheque_number": cheque.cheque_number,
            "bank_name": cheque.bank_name,
            "client_name": client.name,
            "property_name": prop.name,
            "contract_number": contract.contract_number,
            "for_month": cheque.for_month,
            "amount": float(cheque.amount),
            "kind": "Security" if cheque.is_security else "Rent",
            "status": cheque.status,
        })
    return {
        "columns": [
            {"key": "cheque_date", "label": "Cheque Date", "width": 13},
            {"key": "cheque_number", "label": "Cheque No.", "width": 16},
            {"key": "bank_name", "label": "Bank", "width": 22},
            {"key": "client_name", "label": "Client", "width": 28},
            {"key": "property_name", "label": "Property", "width": 24},
            {"key": "contract_number", "label": "Contract", "width": 18},
            {"key": "for_month", "label": "For Month", "width": 12},
            {"key": "amount", "label": "Amount", "width": 14},
            {"key": "kind", "label": "Type", "width": 10},
            {"key": "status", "label": "Status", "width": 12},
        ],
        "rows": rows,
        "meta": {
            "count": len(rows),
            "total": round(sum(r["amount"] for r in rows), 2),
        },
    }


# 9. Property P&L — the New-2026 block
@report(
    "property-pnl",
    "Property Profit & Loss",
    "Money",
    "Per property: rent charged, paid to the landlord, running costs and margin.",
)
def _r_property_pnl(filters: dict) -> dict:
    period = month_start(parse_date(filters.get("month"), date.today()))
    data = property_pnl_service(
        period_month=period,
        property_id=int(filters["property_id"]) if filters.get("property_id") else None,
    )
    rows = [r for r in data["rows"]
            if r["rent_charged"] or r["rent_paid"] or r["expense_total"]]

    # Expense categories become columns, exactly like the workbook block.
    categories = sorted({name for r in rows for name in r["expenses"]})
    columns = [
        {"key": "property_code", "label": "Code", "width": 14},
        {"key": "property_name", "label": "Property", "width": 26},
        {"key": "rent_charged", "label": "Rent Charged", "width": 14},
        {"key": "rent_collected", "label": "Collected", "width": 14},
        {"key": "rent_paid", "label": "Rent Paid", "width": 14},
    ]
    columns += [{"key": f"exp:{c}", "label": c, "width": 16} for c in categories]
    columns += [
        {"key": "expense_total", "label": "Total Costs", "width": 14},
        {"key": "profit", "label": "Profit / Loss", "width": 15},
    ]

    flat = []
    for r in rows:
        row = {k: r[k] for k in
               ("property_code", "property_name", "rent_charged", "rent_collected",
                "rent_paid", "expense_total", "profit")}
        for c in categories:
            row[f"exp:{c}"] = r["expenses"].get(c) or None
        flat.append(row)
    return {
        "columns": columns,
        "rows": flat,
        "meta": {**data["totals"], "count": len(flat), "month": period.isoformat()},
    }


# 10. Company P&L
@report(
    "company-pnl",
    "Company Profit & Loss",
    "Money",
    "The whole company for a month: rental income, direct costs, overhead, net.",
)
def _r_company_pnl(filters: dict) -> dict:
    period = month_start(parse_date(filters.get("month"), date.today()))
    data = property_pnl_service(period_month=period)
    totals = data["totals"]

    by_category = (
        db.session.query(ExpenseCategory.code, ExpenseCategory.name, ExpenseCategory.kind,
                         func.coalesce(func.sum(Expense.amount), 0))
        .join(Expense, Expense.category_id == ExpenseCategory.id)
        .filter(Expense.period_month == period)
        .filter(Expense.status == "posted")
        # RENT_PAID would double-count "Rent paid to landlords" below,
        # already sourced from LandlordPayment/landlord_charges.
        .filter(ExpenseCategory.code != "RENT_PAID")
        .group_by(ExpenseCategory.code, ExpenseCategory.name, ExpenseCategory.kind)
        .order_by(ExpenseCategory.kind, ExpenseCategory.name)
        .all()
    )

    rows = [
        {"line": "Rent charged to clients", "section": "Income",
         "amount": totals.get("rent_charged", 0.0)},
        {"line": "Rent paid to landlords", "section": "Direct cost",
         "amount": -totals.get("rent_paid", 0.0)},
    ]
    for code, name, kind, amount in by_category:
        value = round(float(amount or 0), 2)
        if kind == "income":
            rows.append({"line": name, "section": "Income", "amount": value})
        else:
            rows.append({
                "line": name,
                "section": "Direct cost" if kind == "direct" else "Overhead",
                "amount": -value,
            })
    rows.append({"line": "Net profit", "section": "Result",
                 "amount": totals.get("net_profit", 0.0)})
    return {
        "columns": [
            {"key": "section", "label": "Section", "width": 16},
            {"key": "line", "label": "Line", "width": 36},
            {"key": "amount", "label": "Amount", "width": 16},
        ],
        "rows": rows,
        "meta": {**totals, "month": period.isoformat()},
    }


# 10.5 Monthly P&L comparison
@report(
    "monthly-pnl",
    "Monthly Profit & Loss Comparison",
    "Money",
    "The company P&L across several months side by side, in the accounting "
    "software's own layout — Other Income and Rent Received down to Net Profit.",
)
def _r_monthly_pnl(filters: dict) -> dict:
    end = month_start(parse_date(filters.get("to_month"), date.today()))
    start = parse_date(filters.get("from_month"), None)
    start = month_start(start) if start else add_months(end, -6)
    if start > end:
        start, end = end, start

    months: list[date] = []
    cursor = start
    while cursor <= end:
        months.append(cursor)
        cursor = add_months(cursor, 1)

    # Rent charged (accrual revenue) and rent due to landlords (accrual
    # direct cost) both already exist, fully worked out, in
    # property_pnl()'s company-wide totals — reuse them rather than
    # re-deriving from RentCharge/landlord_charges here.
    monthly_totals = {m: property_pnl_service(period_month=m)["totals"] for m in months}

    other_income = dict(
        db.session.query(Expense.period_month,
                         func.coalesce(func.sum(Expense.amount), 0))
        .join(ExpenseCategory, Expense.category_id == ExpenseCategory.id)
        .filter(ExpenseCategory.code == "OTHER_INCOME")
        .filter(Expense.period_month.in_(months))
        .filter(Expense.status == "posted")
        .group_by(Expense.period_month).all()
    )

    # Every other direct/indirect category, company-wide. RENT_PAID is
    # excluded — it's already carried via rent_due_landlord above.
    cat_rows = (
        db.session.query(Expense.period_month, ExpenseCategory.name, ExpenseCategory.kind,
                         func.coalesce(func.sum(Expense.amount), 0))
        .join(ExpenseCategory, Expense.category_id == ExpenseCategory.id)
        .filter(Expense.period_month.in_(months))
        .filter(Expense.status == "posted")
        .filter(ExpenseCategory.code != "RENT_PAID")
        .filter(ExpenseCategory.kind.in_(("direct", "indirect")))
        .group_by(Expense.period_month, ExpenseCategory.name, ExpenseCategory.kind)
        .all()
    )
    direct_categories: dict[str, dict[date, float]] = {}
    indirect_categories: dict[str, dict[date, float]] = {}
    for period, name, kind, amount in cat_rows:
        bucket = direct_categories if kind == "direct" else indirect_categories
        bucket.setdefault(name, {})[period] = round(float(amount or 0), 2)

    month_keys = [m.strftime("%Y-%m") for m in months]

    def make_row(line: str, section: str, values: dict[date, float]) -> dict:
        row = {"line": line, "section": section}
        for key, m in zip(month_keys, months):
            row[key] = round(float(values.get(m, 0) or 0), 2)
        row["total"] = round(sum(row[k] for k in month_keys), 2)
        return row

    rows = []
    rent_received = {m: monthly_totals[m].get("rent_charged", 0.0) for m in months}
    rows.append(make_row("Other Income", "Revenues", other_income))
    rows.append(make_row("Rent Received", "Revenues", rent_received))
    total_revenues = {m: float(other_income.get(m, 0) or 0) + rent_received[m] for m in months}
    rows.append(make_row("Total Revenues", "Revenues", total_revenues))

    for name in sorted(direct_categories):
        rows.append(make_row(name, "Direct", direct_categories[name]))
    rent_paid_accrual = {m: monthly_totals[m].get("rent_due_landlord", 0.0) for m in months}
    rows.append(make_row("Rent Paid", "Direct", rent_paid_accrual))
    cost_of_sales = {
        m: sum(direct_categories[n].get(m, 0) for n in direct_categories) + rent_paid_accrual[m]
        for m in months
    }
    rows.append(make_row("Cost of Sales", "Direct", cost_of_sales))
    gross_profit = {m: total_revenues[m] - cost_of_sales[m] for m in months}
    rows.append(make_row("Gross Profit / (Loss)", "Result", gross_profit))

    for name in sorted(indirect_categories):
        rows.append(make_row(name, "Indirect", indirect_categories[name]))
    total_indirect = {
        m: sum(indirect_categories[n].get(m, 0) for n in indirect_categories) for m in months
    }
    rows.append(make_row("Total Indirect Expenses", "Indirect", total_indirect))
    net_profit = {m: gross_profit[m] - total_indirect[m] for m in months}
    rows.append(make_row("Net Profit / (Loss)", "Result", net_profit))

    columns = [{"key": "line", "label": "Line", "width": 28}]
    columns += [{"key": k, "label": m.strftime("%b %Y"), "width": 13}
               for k, m in zip(month_keys, months)]
    columns += [{"key": "total", "label": "Total", "width": 15}]

    return {
        "columns": columns,
        "rows": rows,
        "meta": {"count": len(rows), "months": month_keys,
                 "start": start.isoformat(), "end": end.isoformat()},
    }


# 11. Audit Report
@report(
    "audit-trail",
    "Audit Report",
    "Control",
    "Who changed what, when — filterable by module, action, user and date.",
)
def _r_audit(filters: dict) -> dict:
    q = AuditLog.query
    if filters.get("module"):
        q = q.filter(AuditLog.module == filters["module"])
    if filters.get("action"):
        q = q.filter(AuditLog.action == filters["action"])
    if filters.get("user_id"):
        q = q.filter(AuditLog.user_id == int(filters["user_id"]))
    if filters.get("date_from"):
        q = q.filter(AuditLog.created_at >= datetime.combine(
            parse_date(filters["date_from"]), datetime.min.time()))
    if filters.get("date_to"):
        q = q.filter(AuditLog.created_at < datetime.combine(
            parse_date(filters["date_to"]) + timedelta(days=1), datetime.min.time()))

    limit = min(int(filters.get("limit") or 1000), 5000)
    rows = []
    for log in q.order_by(AuditLog.id.desc()).limit(limit).all():
        rows.append({
            "created_at": (log.created_at.isoformat(sep=" ", timespec="seconds")
                           if log.created_at else None),
            "username": log.username,
            "module": log.module,
            "action": log.action,
            "entity": (f"{log.entity_type}#{log.entity_id}"
                       if log.entity_type else None),
            "remarks": log.remarks,
            "ip_address": log.ip_address,
        })
    return {
        "columns": [
            {"key": "created_at", "label": "Time (UTC)", "width": 20},
            {"key": "username", "label": "User", "width": 16},
            {"key": "module", "label": "Module", "width": 16},
            {"key": "action", "label": "Action", "width": 14},
            {"key": "entity", "label": "Entity", "width": 24},
            {"key": "remarks", "label": "Remarks", "width": 34},
            {"key": "ip_address", "label": "IP", "width": 16},
        ],
        "rows": rows,
        "meta": {"count": len(rows), "limit": limit},
    }


# 14. Client risk

@report(
    "client-risk",
    "Client Risk Score",
    "Contracts",
    "Clients ranked by a weighted score of days overdue, bounced cheques "
    "and late payments over the trailing 12 months — who to look at "
    "before renewing.",
)
def _r_client_risk(filters: dict) -> dict:
    rows = client_risk_report()
    return {
        "columns": [
            {"key": "client_code", "label": "Code", "width": 12},
            {"key": "client_name", "label": "Client", "width": 28},
            {"key": "outstanding", "label": "Outstanding", "width": 14},
            {"key": "days_overdue", "label": "Days Overdue", "width": 12},
            {"key": "bounced_cheques_12m", "label": "Bounced Cheques (12m)", "width": 16},
            {"key": "late_payments_12m", "label": "Late Payments (12m)", "width": 16},
            {"key": "score", "label": "Score", "width": 10},
            {"key": "tier", "label": "Tier", "width": 10},
        ],
        "rows": rows,
        "meta": {"count": len(rows),
                 "high": sum(1 for r in rows if r["tier"] == "high"),
                 "medium": sum(1 for r in rows if r["tier"] == "medium")},
    }
