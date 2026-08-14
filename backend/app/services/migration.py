"""One-time migration off `1-GreenTech Master File 2026.xlsm`.

The workbook is a spreadsheet kept by hand for years, so it is *mostly*
regular and occasionally not. This module reads it into a staged plan
and **writes nothing**. Everything it could not understand is reported
as a problem for a person to resolve before the plan is committed.

Shape of the New-2026 sheet, per property:

    <integer>                              block index
    AL SAFA ENGENERING STREET 10 …         property title
    Client | LAND LOARD | Expire | MODE | Rooms | OP Dec-2025 | Jan … Jul
    …tenant rows, cash section…
    Cash Total
    …tenant rows, cheque section…
    Cheque Total
    EMPTY ROOMS                            count in the Rooms column
    TOTAL                                  total rooms in the Rooms column
    Rent Paid                              paid to the landlord, per month
    Sewage Removal And Cleaning            direct costs, per month
    Eelctricity and Water
    Maint & Cleaning
    Profit Or Loss
    Fund Flow

A tenant row carries the whole year across the month columns, so the
rent in each month is known. The migration turns that into a contract
starting at the first month with a figure, plus a dated `rent_change`
amendment wherever the figure moves — which is what those changing
columns actually *were*. Reproducing them as amendments rather than
flattening to a single rent is what lets the parallel run compare the
app's month-by-month output against the sheet it came from.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta

from openpyxl import load_workbook


SHEET_MAIN = "New-2026"
SHEET_LANDLORD = "LANDLOARD"
SHEET_EMPTY = "Empty Rooms"

# Column letters are fixed by the sheet's layout; month columns are read
# from the header row rather than assumed, so adding August won't break it.
COL_CLIENT, COL_PROPERTY, COL_EXPIRE, COL_MODE, COL_ROOMS, COL_OPENING = 1, 2, 3, 4, 5, 6

# Rows that carry structure rather than a tenant.
STRUCTURAL = {
    "cash total", "cheque total", "online total", "bank total",
    "total", "grand total", "empty rooms", "empty room",
    "profit or loss", "fund flow",
}
# Rows that carry money out rather than in. `rent paid` is the landlord
# payment; the rest are the property's running costs.
LANDLORD_PAYMENT_LABEL = "rent paid"
# Sheet label -> the seeded expense category *code*. Codes rather than
# names because a name can be re-worded in Settings without warning; the
# code is what the rest of the system keys on.
EXPENSE_LABELS = {
    "sewage removal and cleaning": "SEWAGE",
    "eelctricity and water": "ELECTRICITY_CAMP",       # the sheet's own spelling
    "electricity and water": "ELECTRICITY_CAMP",
    "maint & cleaning": "MAINT_CLEANING",
    "maintenance & cleaning": "MAINT_CLEANING",
    "other extra expense": "GENERAL_CAMP",
    "direct operating expense": None,      # a computed roll-up, not a cost
}

PAYMENT_MODES = {"cash": "cash", "cheque": "cheque", "online": "online",
                 "bank": "online", "transfer": "online"}

# Statuses that appear in the Expire column instead of a date. They mean
# the tenancy has ended; only the wording differs.
ENDED_STATUSES = {"cancelled", "cancel", "transferred", "transfer", "vacated",
                  "closed", "shifted"}
CANCELLED = "cancelled"

# Where the per-property blocks stop and the company roll-up begins.
TAIL_MARKERS = ("cancelled & pending", "cancelled and pending")


def _payment_mode(raw: str) -> tuple[str, str | None]:
    """Read the MODE column, which is free text rather than a code.

    Real values include `M.CHEQUE` (monthly cheque) and `ONLINE AG.CHQ`
    (paid online, with an agreement cheque held). Falling back to cash on
    those would be worse than a guess — it would tell the portal not to
    expect a cheque book. So look for the payment word anywhere in the
    text, and only report a problem when nothing is recognisable.
    """
    text = _norm(raw)
    if not text:
        return "cash", None
    exact = PAYMENT_MODES.get(text)
    if exact:
        return exact, None
    # `online ag.chq` is online money with a cheque on file — the money
    # matters more than the paper, so online wins where both appear.
    if "online" in text or "transfer" in text or "bank" in text:
        return "online", None
    if "chq" in text or "cheque" in text or "check" in text:
        return "cheque", None
    if "cash" in text:
        return "cash", None
    return "cash", f"unknown payment mode {text!r}, treating as cash"


def _norm(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def _clean(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _number(value) -> float | None:
    if value is None or isinstance(value, (datetime, date)):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _as_date(value):
    """Return (date | None, problem | None).

    The sheet contains dates that do not exist — `31/06/2026` — because
    they were typed as text. Those must surface, not be rounded into
    something plausible.
    """
    if value is None or value == "":
        return None, None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    text = str(value).strip()
    if _norm(text) in ENDED_STATUSES:
        return None, None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date(), None
        except ValueError:
            continue
    return None, f"{text!r} is not a usable date"


@dataclass
class Problem:
    where: str
    detail: str
    severity: str = "warning"     # warning | blocker

    def to_dict(self) -> dict:
        return {"where": self.where, "detail": self.detail, "severity": self.severity}


@dataclass
class TenantRow:
    row: int
    name: str
    payment_mode: str
    rooms: int
    unit_hint: str | None            # 'SHOP' / 'STORE' when Rooms held a word
    expiry: date | None
    is_cancelled: bool
    opening_balance: float
    monthly: dict[str, float]        # 'YYYY-MM-01' -> rent for that month

    def rent_timeline(self) -> list[tuple[str, float]]:
        return sorted(self.monthly.items())

    def first_month(self) -> str | None:
        live = [m for m, v in self.rent_timeline() if v]
        return live[0] if live else None

    def last_month(self) -> str | None:
        live = [m for m, v in self.rent_timeline() if v]
        return live[-1] if live else None

    def to_dict(self) -> dict:
        return {
            "row": self.row, "name": self.name, "payment_mode": self.payment_mode,
            "rooms": self.rooms, "unit_hint": self.unit_hint,
            "expiry": self.expiry.isoformat() if self.expiry else None,
            "is_cancelled": self.is_cancelled,
            "opening_balance": self.opening_balance,
            "monthly": self.monthly,
            "first_month": self.first_month(), "last_month": self.last_month(),
        }


@dataclass
class PropertyBlock:
    index: int
    title_row: int
    title: str
    tenants: list[TenantRow] = field(default_factory=list)
    total_rooms: int | None = None
    empty_rooms: int | None = None
    total: dict[str, float] = field(default_factory=dict)   # the TOTAL row's months
    rent_paid: dict[str, float] = field(default_factory=dict)
    expenses: dict[str, dict[str, float]] = field(default_factory=dict)
    profit: dict[str, float] = field(default_factory=dict)
    landlord_row: int | None = None
    landlord_match: str | None = None      # exact | fuzzy | none

    def to_dict(self) -> dict:
        return {
            "index": self.index, "title_row": self.title_row, "title": self.title,
            "tenants": [t.to_dict() for t in self.tenants],
            "total_rooms": self.total_rooms, "empty_rooms": self.empty_rooms,
            "total": self.total,
            "rent_paid": self.rent_paid, "expenses": self.expenses,
            "profit": self.profit,
            "landlord_row": self.landlord_row, "landlord_match": self.landlord_match,
        }


@dataclass
class LandlordRow:
    row: int
    title: str
    english: str
    arabic: str
    location: str
    start: date | None
    expiry: date | None
    rent: float | None
    rooms: int | None
    stores: int | None
    security_cheque: float | None
    opening_balance: float | None

    def to_dict(self) -> dict:
        return {
            "row": self.row, "title": self.title, "english": self.english,
            "arabic": self.arabic, "location": self.location,
            "start": self.start.isoformat() if self.start else None,
            "expiry": self.expiry.isoformat() if self.expiry else None,
            "rent": self.rent, "rooms": self.rooms, "stores": self.stores,
            "security_cheque": self.security_cheque,
            "opening_balance": self.opening_balance,
        }


# ----------------------------------------------------------------------
# Fuzzy matching between a block title and a landlord row
# ----------------------------------------------------------------------

_NOISE = re.compile(
    r"\b(street|st|no|nos|rooms?|room|store|stores|shop|shops|and|the|camp|"
    r"aggrement|agreement|deposit|chq|qr)\b")


def _fingerprint(title: str) -> set[str]:
    """Words that actually identify a property: the owner's name and any
    street number, with the boilerplate stripped out."""
    text = _norm(title)
    text = re.sub(r"[\[\]\(\),.\-+#:]", " ", text)
    text = _NOISE.sub(" ", text)
    return {w for w in text.split() if len(w) > 1}


def _similarity(a: str, b: str) -> float:
    fa, fb = _fingerprint(a), _fingerprint(b)
    if not fa or not fb:
        return 0.0
    return len(fa & fb) / len(fa | fb)


# ----------------------------------------------------------------------
# Parsing
# ----------------------------------------------------------------------

def _month_columns(ws, header_row: int) -> dict[int, str]:
    """Map column index -> 'YYYY-MM-01' from the header row's dates."""
    out = {}
    for col in range(COL_OPENING + 1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if isinstance(value, (datetime, date)):
            day = value.date() if isinstance(value, datetime) else value
            out[col] = day.replace(day=1).isoformat()
    return out


def _parse_tenant(ws, row: int, months: dict[int, str],
                  problems: list[Problem]) -> TenantRow | None:
    name = _clean(ws.cell(row, COL_CLIENT).value)
    if not name:
        return None

    raw_expire = ws.cell(row, COL_EXPIRE).value
    is_cancelled = _norm(raw_expire) in ENDED_STATUSES
    expiry, date_problem = _as_date(raw_expire)
    if date_problem:
        problems.append(Problem(f"New-2026!C{row}",
                                f"{name}: {date_problem}", "warning"))

    mode, mode_problem = _payment_mode(ws.cell(row, COL_MODE).value)
    if mode_problem:
        problems.append(Problem(f"New-2026!D{row}", f"{name}: {mode_problem}"))

    rooms_raw = ws.cell(row, COL_ROOMS).value
    rooms_number = _number(rooms_raw)
    unit_hint = None
    if rooms_number is None and rooms_raw:
        # 'SHOP' / 'STORE' — one unit of that kind rather than a count.
        unit_hint = _clean(rooms_raw).upper()
        rooms = 1
    else:
        rooms = int(rooms_number or 0)

    monthly = {}
    for col, month in months.items():
        value = _number(ws.cell(row, col).value)
        if value is not None:
            monthly[month] = round(value, 2)

    return TenantRow(
        row=row, name=name, payment_mode=mode, rooms=rooms, unit_hint=unit_hint,
        expiry=expiry, is_cancelled=is_cancelled,
        opening_balance=round(_number(ws.cell(row, COL_OPENING).value) or 0, 2),
        monthly=monthly,
    )


def _tail_row(ws) -> int:
    """First row of the `Cancelled & Pending Receivable` tail.

    Everything below it is company-level roll-up, not a property block —
    and the last block would otherwise run on and swallow it, quietly
    attributing the whole company's costs to one building.
    """
    for row in range(1, ws.max_row + 1):
        if _norm(ws.cell(row, COL_CLIENT).value).startswith(TAIL_MARKERS):
            return row
    return ws.max_row + 1


def _parse_blocks(ws, problems: list[Problem]) -> list[PropertyBlock]:
    blocks: list[PropertyBlock] = []
    limit = _tail_row(ws)
    row = 1
    while row < limit:
        marker = ws.cell(row, COL_CLIENT).value
        is_index = isinstance(marker, (int, float)) and not isinstance(marker, bool)
        if not is_index:
            row += 1
            continue

        title = _clean(ws.cell(row + 1, COL_CLIENT).value)
        header_row = row + 2
        if not title or _norm(ws.cell(header_row, COL_CLIENT).value) != "client":
            row += 1
            continue

        months = _month_columns(ws, header_row)
        if not months:
            problems.append(Problem(f"New-2026!{header_row}",
                                    f"{title}: no month columns found", "blocker"))
        block = PropertyBlock(index=int(marker), title_row=row + 1, title=title)

        row = header_row + 1
        while row < limit:
            label = _norm(ws.cell(row, COL_CLIENT).value)
            nxt = ws.cell(row, COL_CLIENT).value
            if isinstance(nxt, (int, float)) and not isinstance(nxt, bool):
                break                                  # next block's index
            monthly = {m: _number(ws.cell(row, c).value)
                       for c, m in months.items()}
            has_money = any(v for v in monthly.values())

            if label in STRUCTURAL:
                rooms = _number(ws.cell(row, COL_ROOMS).value)
                if label in ("empty rooms", "empty room"):
                    block.empty_rooms = int(rooms or 0)
                elif label in ("total", "grand total"):
                    block.total_rooms = int(rooms or 0)
                    block.total = {m: v for m, v in monthly.items() if v is not None}
                elif label == "profit or loss":
                    block.profit = {m: v for m, v in monthly.items() if v is not None}
            elif label == LANDLORD_PAYMENT_LABEL:
                block.rent_paid = {m: v for m, v in monthly.items() if v}
            elif label in EXPENSE_LABELS:
                category = EXPENSE_LABELS[label]
                if category:
                    block.expenses[category] = {m: v for m, v in monthly.items() if v}
            elif label:
                tenant = _parse_tenant(ws, row, months, problems)
                if tenant is not None:
                    block.tenants.append(tenant)
            elif has_money:
                problems.append(Problem(
                    f"New-2026!A{row}",
                    f"{block.title}: a row carries money but has no name", "warning"))
            row += 1

        blocks.append(block)
    return blocks


def _parse_landlords(ws, problems: list[Problem]) -> list[LandlordRow]:
    out = []
    for row in range(2, ws.max_row + 1):
        title = _clean(ws.cell(row, 2).value)
        english = _clean(ws.cell(row, 3).value)
        if not title:
            continue
        if _norm(ws.cell(row, 1).value) in ("sl", "s.no", "sl "):
            break                                       # the second table starts
        start, start_problem = _as_date(ws.cell(row, 6).value)
        expiry, expiry_problem = _as_date(ws.cell(row, 7).value)
        for problem, column in ((start_problem, "F"), (expiry_problem, "G")):
            if problem:
                problems.append(Problem(
                    f"{SHEET_LANDLORD}!{column}{row}",
                    f"{english or title}: {problem}", "warning"))
        out.append(LandlordRow(
            row=row, title=title, english=english or title,
            arabic=_clean(ws.cell(row, 4).value),
            location=_clean(ws.cell(row, 5).value),
            start=start, expiry=expiry,
            rent=_number(ws.cell(row, 8).value),
            rooms=int(_number(ws.cell(row, 9).value) or 0) or None,
            stores=int(_number(ws.cell(row, 10).value) or 0) or None,
            security_cheque=_number(ws.cell(row, 11).value),
            opening_balance=_number(ws.cell(row, 12).value),
        ))
    return out


def _parse_receivables(ws, months: dict[int, str]) -> list[dict]:
    """The `Cancelled & Pending Receivable` tail — debts still owed by
    tenants who have already gone."""
    out = []
    start = None
    for row in range(1, ws.max_row + 1):
        if _norm(ws.cell(row, COL_CLIENT).value).startswith("cancelled & pending"):
            start = row + 2
            break
    if start is None:
        return out
    for row in range(start, ws.max_row + 1):
        name = _clean(ws.cell(row, COL_CLIENT).value)
        if not name:
            continue
        if _norm(name) in STRUCTURAL or _norm(name) == "client":
            continue
        opening = _number(ws.cell(row, COL_OPENING).value)
        if not opening:
            continue
        out.append({
            "row": row, "name": name,
            "property_title": _clean(ws.cell(row, COL_PROPERTY).value) or None,
            "payment_mode": PAYMENT_MODES.get(
                _norm(ws.cell(row, COL_MODE).value), "cash"),
            "opening_balance": round(opening, 2),
        })
    return out


def _parse_company_totals(ws, months: dict[int, str]) -> dict:
    """The roll-up at the foot of the sheet — the figures the parallel
    run has to reproduce."""
    wanted = {
        "opening balance 01-01-2026": "opening_balance",
        "total rent received": "rent_received",
        "rent paid": "rent_paid",
        "other income": "other_income",
        "sewage removal and cleaning": "sewage",
        "eelctricity and water": "electricity_water",
        "maint & cleaning": "maintenance",
        "direct operating expense": "direct_operating_expense",
        "gross profit or loss": "gross_profit",
    }
    totals: dict[str, dict[str, float]] = {}
    for row in range(1, ws.max_row + 1):
        label = _norm(ws.cell(row, COL_PROPERTY).value)   # this block labels in B
        key = wanted.get(label)
        if not key or key in totals:
            continue
        totals[key] = {m: v for c, m in months.items()
                       if (v := _number(ws.cell(row, c).value)) is not None}
    # The cash / cheque split at the very bottom is labelled in column A.
    for row in range(ws.max_row, max(ws.max_row - 12, 1), -1):
        label = _norm(ws.cell(row, COL_CLIENT).value)
        if label in ("cash total", "cheque total", "total"):
            key = {"cash total": "cash_total", "cheque total": "cheque_total",
                   "total": "grand_total"}[label]
            totals.setdefault(key, {m: v for c, m in months.items()
                                    if (v := _number(ws.cell(row, c).value)) is not None})
    return totals


def _parse_empty_rooms(wb) -> list[dict]:
    if SHEET_EMPTY not in wb.sheetnames:
        return []
    ws = wb[SHEET_EMPTY]
    out = []
    for row in range(2, ws.max_row + 1):
        landlord = _clean(ws.cell(row, 2).value)
        if not landlord:
            continue
        out.append({
            "landlord": landlord,
            "location": _clean(ws.cell(row, 3).value),
            "total_rooms": int(_number(ws.cell(row, 4).value) or 0),
            "occupied": int(_number(ws.cell(row, 5).value) or 0),
            "empty": int(_number(ws.cell(row, 6).value) or 0),
            "stores": int(_number(ws.cell(row, 7).value) or 0),
            "stores_occupied": int(_number(ws.cell(row, 8).value) or 0),
        })
    return out


def parse_workbook(path: str) -> dict:
    """Read the whole workbook into a staged plan. Writes nothing."""
    problems: list[Problem] = []
    wb = load_workbook(path, data_only=True, read_only=False)

    for required in (SHEET_MAIN, SHEET_LANDLORD):
        if required not in wb.sheetnames:
            return {"ok": False, "problems": [
                Problem(path, f"Sheet {required!r} is missing", "blocker").to_dict()]}

    ws = wb[SHEET_MAIN]
    blocks = _parse_blocks(ws, problems)
    landlords = _parse_landlords(wb[SHEET_LANDLORD], problems)

    # Month columns for the sheet-level sections, taken from the first
    # block's header (they are the same throughout).
    header_row = blocks[0].title_row + 1 if blocks else 3
    months = _month_columns(ws, header_row)

    # --- join each block to its landlord
    by_title = {_norm(l.title): l for l in landlords}
    for block in blocks:
        exact = by_title.get(_norm(block.title))
        if exact is not None:
            block.landlord_row, block.landlord_match = exact.row, "exact"
            continue
        best, score = None, 0.0
        for candidate in landlords:
            value = _similarity(block.title, candidate.title)
            if value > score:
                best, score = candidate, value
        if best is not None and score >= 0.45:
            block.landlord_row, block.landlord_match = best.row, "fuzzy"
            problems.append(Problem(
                f"New-2026!A{block.title_row}",
                f"{block.title!r} was matched to landlord {best.english!r} "
                f"by similarity ({score:.0%}) — confirm before committing"))
        else:
            block.landlord_match = "none"
            problems.append(Problem(
                f"New-2026!A{block.title_row}",
                f"{block.title!r} has no landlord row in {SHEET_LANDLORD}; "
                "a landlord will be created from the block title", "warning"))

    receivables = _parse_receivables(ws, months)
    totals = _parse_company_totals(ws, months)

    tenant_count = sum(len(b.tenants) for b in blocks)
    if not blocks:
        problems.append(Problem(SHEET_MAIN, "No property blocks found", "blocker"))

    return {
        "ok": not any(p.severity == "blocker" for p in problems),
        "source": path,
        "months": sorted(set(months.values())),
        "blocks": [b.to_dict() for b in blocks],
        "landlords": [l.to_dict() for l in landlords],
        "receivables": receivables,
        "company_totals": totals,
        "empty_rooms": _parse_empty_rooms(wb),
        "problems": [p.to_dict() for p in problems],
        "summary": {
            "properties": len(blocks),
            "landlords": len(landlords),
            "tenants": tenant_count,
            "receivables": len(receivables),
            "months": len(months),
            "exact_landlord_matches": sum(1 for b in blocks if b.landlord_match == "exact"),
            "fuzzy_landlord_matches": sum(1 for b in blocks if b.landlord_match == "fuzzy"),
            "unmatched_properties": sum(1 for b in blocks if b.landlord_match == "none"),
            "blockers": sum(1 for p in problems if p.severity == "blocker"),
            "warnings": sum(1 for p in problems if p.severity == "warning"),
        },
    }


# ======================================================================
# Committing the plan
# ======================================================================

STORE_WORDS = ("store", "shop", "cafteria", "cafeteria", "supermarket", "kitchen")


def _property_name(title: str) -> str:
    """A short, recognisable name for the property.

    The workbook titles carry the owner, the street and the room count in
    one string. The street is what the staff actually call the building
    ("ST-10"), so lead with that where it can be found.
    """
    text = _clean(title)
    match = re.search(r"(?:street|st)\.?\s*(?:no\.?)?\s*[-–]?\s*(\d{1,3})", text, re.I)
    street = f"ST-{match.group(1)}" if match else None
    # A qualifier distinguishes several buildings on one street — the
    # staff say "38 Kumar" and "38 China", never the owner's full name.
    qualifier = None
    for word in ("KUMAR", "COMPUTER", "HYPER", "CHINA", "PARIS", "MILKA",
                 "MILANO", "MAHARIB", "BIRKAT", "SADD TRUCK", "YTC"):
        if re.search(rf"\b{word}\b", text, re.I):
            qualifier = word.upper()
            break
    if street and qualifier:
        return f"{street} {qualifier}"
    if street:
        return street
    if qualifier:
        return qualifier
    # No street number anywhere: fall back to the owner's name, trimmed of
    # the room/store boilerplate so it reads as a place, not a sentence.
    trimmed = re.split(r"\s*[-–(\[]", text)[0]
    trimmed = re.sub(r"\b(street|st)\b.*$", "", trimmed, flags=re.I).strip()
    return (trimmed or text)[:60]


def _property_type(title: str, rooms: int, stores: int) -> str:
    lowered = title.lower()
    if "camp" in lowered:
        return "labour_camp"
    if rooms and stores:
        return "building_with_store"
    if rooms:
        return "full_building"
    if stores:
        return "store"
    return "mixed_use"


def _floor_split(units: int, per_floor: int = 50) -> tuple[int, int]:
    """Spread `units` over enough floors to stay inside the layout
    builder's limits. ST-47 has 144 rooms; one floor cannot hold them."""
    units = max(int(units or 0), 1)
    floors = max(1, -(-units // per_floor))          # ceil
    return floors, -(-units // floors)               # ceil again, evened out


def build_commit_plan(parsed: dict, *, overrides: dict | None = None) -> dict:
    """Turn a parsed workbook into the exact set of writes, and check the
    arithmetic before anything is written.

    `overrides` lets the review screen correct what the parser guessed:
    ``{"property_names": {block_index: name},
       "landlord_names": {block_index: name},
       "skip_blocks": [index, ...],
       "skip_tenants": {block_index: [row, ...]}}``

    `skip_tenants` drops specific tenant rows from a block before
    contracts are built from it — for a row that turns out to be a
    manual note rather than a real tenancy (e.g. a one-off rent
    adjustment line), not a whole property.
    """
    overrides = overrides or {}
    names = overrides.get("property_names") or {}
    landlord_names = overrides.get("landlord_names") or {}
    skip = set(overrides.get("skip_blocks") or [])
    skip_tenants_raw = overrides.get("skip_tenants") or {}
    skip_tenants = {
        int(k): set(v) for k, v in skip_tenants_raw.items()
    }
    used_names: dict[str, int] = {}

    landlords_by_row = {l["row"]: l for l in parsed["landlords"]}
    problems = list(parsed.get("problems", []))
    plan_blocks = []

    for block in parsed["blocks"]:
        if block["index"] in skip:
            continue
        drop_rows = skip_tenants.get(block["index"]) or set()
        if drop_rows:
            block = {**block, "tenants": [t for t in block["tenants"]
                                           if t["row"] not in drop_rows]}
        landlord = landlords_by_row.get(block["landlord_row"])
        landlord_name = (landlord_names.get(str(block["index"]))
                         or landlord_names.get(block["index"])
                         or (landlord["english"] if landlord else _clean(block["title"])[:60]))
        property_name = (names.get(str(block["index"]))
                         or names.get(block["index"]))
        if not property_name:
            property_name = _property_name(block["title"])
            # Three different owners have buildings on street 38, so the
            # street alone is not a name. Where it repeats, add the
            # owner's first word — merging two buildings into one would
            # double-let their rooms.
            if property_name in used_names:
                surname = _clean(landlord_name).split()[0] if landlord_name else ""
                candidate = f"{property_name} {surname}".strip()
                while candidate in used_names:
                    used_names[candidate] = used_names.get(candidate, 0) + 1
                    candidate = f"{property_name} {surname} {used_names[candidate]}".strip()
                property_name = candidate
        used_names[property_name] = used_names.get(property_name, 0) + 1

        rooms = block["total_rooms"] or (landlord["rooms"] if landlord else 0) or 0
        stores = (landlord["stores"] if landlord else 0) or 0

        # How many units the tenancies actually need, so the property is
        # built big enough to hold them even where the sheet's own room
        # count disagrees with the sum of its rows.
        room_demand = sum(t["rooms"] for t in block["tenants"]
                          if not _demand_is_store(t))
        store_demand = sum(max(t["rooms"], 1) for t in block["tenants"]
                           if _demand_is_store(t))
        if room_demand > rooms:
            problems.append(Problem(
                f"New-2026!A{block['title_row']}",
                f"{property_name}: the rows claim {room_demand} rooms but the "
                f"block's TOTAL says {rooms}; building {room_demand} so every "
                "tenancy fits").to_dict())
            rooms = room_demand
        if store_demand > stores:
            stores = store_demand

        plan_blocks.append({
            "index": block["index"],
            "title": block["title"],
            "property_name": property_name,
            "property_type": _property_type(block["title"], rooms, stores),
            "landlord_name": landlord_name,
            "landlord_name_ar": landlord["arabic"] if landlord else None,
            "landlord_match": block["landlord_match"],
            "agreement": {
                "start": landlord["start"] if landlord else None,
                "expiry": landlord["expiry"] if landlord else None,
                "monthly_rent": landlord["rent"] if landlord else None,
                "security_deposit": landlord["security_cheque"] if landlord else None,
            } if landlord else None,
            "rooms": rooms,
            "stores": stores,
            "empty_rooms": block["empty_rooms"],
            "tenants": block["tenants"],
            "rent_paid": block["rent_paid"],
            "expenses": block["expenses"],
        })

    contracts = sum(len(b["tenants"]) for b in plan_blocks)
    return {
        "blocks": plan_blocks,
        "receivables": parsed["receivables"],
        "months": parsed["months"],
        "company_totals": parsed["company_totals"],
        "problems": problems,
        "summary": {
            "properties": len(plan_blocks),
            "units": sum(b["rooms"] + b["stores"] for b in plan_blocks),
            "contracts": contracts,
            "receivables": len(parsed["receivables"]),
            "landlord_payments": sum(len(b["rent_paid"]) for b in plan_blocks),
            "expense_rows": sum(len(months) for b in plan_blocks
                                for months in b["expenses"].values()),
            "blockers": sum(1 for p in problems if p["severity"] == "blocker"),
            "warnings": sum(1 for p in problems if p["severity"] == "warning"),
        },
    }


def _demand_is_store(tenant: dict) -> bool:
    """Whether this tenancy occupies a store rather than rooms.

    Only the Rooms column decides. It says `SHOP` or `STORE` where the
    letting is commercial, and a number where it is rooms — that is the
    sheet's own signal. Reading the tenant's *name* instead looks
    tempting and is wrong: "MILANO KITCHEN AND CLOSET" is a company that
    rents rooms, and guessing from its name invented twenty stores in a
    building that has one.
    """
    hint = (tenant.get("unit_hint") or "").lower()
    return bool(hint) and any(w in hint for w in STORE_WORDS)


class MigrationError(RuntimeError):
    pass


def _add_unit(db, Unit, prop, existing, unit_type: str, actor):
    """Add one more unit to a property whose stated room count fell short.

    Numbered after the highest existing unit on the same floor so it
    reads naturally in the floor map.
    """
    floor_id = existing[0].floor_id if existing else None
    if floor_id is None:
        raise MigrationError(f"{prop.name} has no floor to add a unit to")
    prefix = "STORE-" if unit_type == "store" else ""
    taken_numbers = {u.unit_number for u in existing}
    seq = len(existing) + 1
    while f"{prefix}{seq}" in taken_numbers:
        seq += 1
    unit = Unit(property_id=prop.id, floor_id=floor_id,
                unit_number=f"{prefix}{seq}", unit_type=unit_type,
                occupancy_status="empty",
                created_by=actor.id, updated_by=actor.id)
    db.session.add(unit)
    db.session.flush()
    return unit


def _iso(value) -> date | None:
    if not value:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    try:
        return datetime.fromisoformat(str(value)).date()
    except ValueError:
        return None


def _month_end(day: date) -> date:
    nxt = day.replace(day=28) + timedelta(days=4)
    return nxt - timedelta(days=nxt.day)


def commit_plan(plan: dict, *, actor, year_end: date | None = None) -> dict:
    """Write the plan. One transaction — the caller commits.

    Idempotent by name: a landlord, client or property that already
    exists is reused rather than duplicated, and a tenancy that already
    exists for the same client, property and start date is left alone.
    That means a half-finished migration can simply be re-run.
    """
    from ..extensions import db
    from ..models import (
        Client, ClientContract, Expense, ExpenseCategory, Landlord,
        LandlordPayment, Property, PropertyAgreement, Unit,
    )
    from . import codes
    from . import contracts as contract_service
    from . import expenses as expense_service
    from . import layout as layout_service
    from . import rent as rent_service

    year_end = year_end or date(2026, 12, 31)
    created = {"landlords": 0, "properties": 0, "units": 0, "clients": 0,
               "contracts": 0, "amendments": 0, "cancellations": 0,
               "landlord_payments": 0, "expenses": 0, "receivables": 0}
    reused = {"landlords": 0, "properties": 0, "clients": 0, "contracts": 0}
    notes: list[str] = []

    def get_landlord(name: str, name_ar: str | None) -> Landlord:
        existing = Landlord.query.filter(
            db.func.lower(Landlord.name) == name.lower()).first()
        if existing:
            reused["landlords"] += 1
            return existing
        row = Landlord(code=codes.next_code(Landlord, codes.prefix_for("landlord")),
                       name=name, name_ar=name_ar or None,
                       created_by=actor.id, updated_by=actor.id)
        db.session.add(row)
        db.session.flush()
        created["landlords"] += 1
        return row

    def get_client(name: str) -> Client:
        existing = Client.query.filter(
            db.func.lower(Client.name) == name.lower()).first()
        if existing:
            reused["clients"] += 1
            return existing
        row = Client(code=codes.next_code(Client, codes.prefix_for("client")),
                     name=name, created_by=actor.id, updated_by=actor.id)
        db.session.add(row)
        db.session.flush()
        created["clients"] += 1
        return row

    categories = {c.code: c for c in ExpenseCategory.query.all()}

    for block in plan["blocks"]:
        landlord = get_landlord(block["landlord_name"], block.get("landlord_name_ar"))

        prop = Property.query.filter(
            db.func.lower(Property.name) == block["property_name"].lower()).first()
        if prop is None:
            prop = Property(code=codes.next_code(Property, codes.prefix_for("property")),
                            name=block["property_name"],
                            property_type=block["property_type"],
                            landlord_id=landlord.id, status="active",
                            created_by=actor.id, updated_by=actor.id)
            db.session.add(prop)
            db.session.flush()
            created["properties"] += 1
        else:
            reused["properties"] += 1

        # --- structure. One floor holding every room keeps the numbers
        # the same as the sheet, which lists rooms without floors. The
        # operator renumbers them properly during the guided review.
        units = Unit.query.filter_by(property_id=prop.id).order_by(Unit.id).all()
        if not units:
            floors, per_floor = _floor_split(block["rooms"])
            layout_service.generate_structure(
                prop, floors=floors, units_per_floor=per_floor, actor=actor)
            units = Unit.query.filter_by(property_id=prop.id).order_by(Unit.id).all()
            created["units"] += len(units)
            if block["stores"]:
                floor_id = units[0].floor_id
                for n in range(block["stores"]):
                    store = Unit(property_id=prop.id, floor_id=floor_id,
                                 unit_number=f"STORE-{n + 1}", unit_type="store",
                                 occupancy_status="empty",
                                 created_by=actor.id, updated_by=actor.id)
                    db.session.add(store)
                    created["units"] += 1
                db.session.flush()
                units = Unit.query.filter_by(property_id=prop.id).order_by(Unit.id).all()

        # Only offer units nothing already holds. Without this a re-run,
        # or two blocks that resolved to the same building, would try to
        # let the same room twice — and the contract service would (
        # rightly) refuse, aborting the whole migration.
        from ..models import ContractUnit
        taken = {
            row.unit_id for row in
            db.session.query(ContractUnit.unit_id)
            .join(ClientContract, ContractUnit.contract_id == ClientContract.id)
            .filter(ClientContract.property_id == prop.id)
            .filter(ContractUnit.to_date.is_(None))
            .all()
        }
        free = [u for u in units if u.id not in taken]
        rooms = [u for u in free if u.unit_type != "store"]
        stores = [u for u in free if u.unit_type == "store"]
        next_room, next_store = 0, 0

        # --- landlord agreement
        agreement = block.get("agreement")
        if agreement and agreement.get("start") and not prop.agreements:
            db.session.add(PropertyAgreement(
                property_id=prop.id, landlord_id=landlord.id,
                start_date=_iso(agreement["start"]),
                expiry_date=_iso(agreement["expiry"]) or year_end,
                monthly_rent=agreement.get("monthly_rent"),
                security_deposit=agreement.get("security_deposit"),
                created_by=actor.id, updated_by=actor.id))
            db.session.flush()

        # --- tenancies
        for tenant in block["tenants"]:
            start = _iso(tenant["first_month"])
            if start is None:
                # No rent in any month: a name with an opening balance
                # only. Carried as a receivable further down.
                if tenant["opening_balance"]:
                    notes.append(f"{tenant['name']}: opening balance only, no rent "
                                 "in any month — migrated as a receivable")
                continue

            client = get_client(tenant["name"])
            existing = ClientContract.query.filter_by(
                client_id=client.id, property_id=prop.id, start_date=start).first()
            if existing is not None:
                reused["contracts"] += 1
                continue

            wants_store = _demand_is_store(tenant)
            count = max(tenant["rooms"], 1)
            if wants_store:
                take = stores[next_store:next_store + count]
                next_store += len(take)
                if not take:                       # ran out of stores; use rooms
                    take = rooms[next_room:next_room + count]
                    next_room += len(take)
            else:
                take = rooms[next_room:next_room + count]
                next_room += len(take)

            if not take:
                # The sheet's room count disagrees with its own rows —
                # it is a hand-kept total and drifts. The tenancy is the
                # harder evidence: somebody is paying rent for a space,
                # so the space exists. Add it rather than drop the rent
                # on the floor, and say so.
                extra = _add_unit(db, Unit, prop, units,
                                  "store" if wants_store else "room", actor)
                take = [extra]
                units.append(extra)
                notes.append(
                    f"{block['property_name']}: the block's room count had no "
                    f"space left for {tenant['name']}, so unit "
                    f"{extra.unit_number} was added to hold the tenancy")

            timeline = sorted(tenant["monthly"].items())
            opening_rent = next((v for _, v in timeline if v), 0)

            # A row that stops paying part-way through the year is a
            # tenancy that ended, whatever the Expire column still says.
            # Without this the portal keeps billing a tenant the sheet
            # stopped billing, and every month after that disagrees.
            last_paid = _iso(tenant["last_month"])
            final_month = _iso(plan["months"][-1]) if plan.get("months") else None
            stops_early = bool(last_paid and final_month and last_paid < final_month)

            # The rent columns are the truth about occupancy; the Expire
            # column is an administrative note and is sometimes stale.
            # One tenancy here carries an expiry of Jan-2025 while the
            # sheet goes on billing it every month of 2026 — honouring
            # that date would end the tenancy on its first day and lose
            # the rent. So a stated expiry can extend the term, never
            # cut it short of what was actually billed.
            billed_to = _month_end(last_paid) if last_paid else start
            stated = _iso(tenant["expiry"])
            expiry = max(billed_to, start)
            if stated and stated > expiry:
                expiry = stated
            elif stated and stated < billed_to:
                notes.append(
                    f"{tenant['name']} ({block['property_name']}): the Expire "
                    f"column says {stated.isoformat()} but rent is billed to "
                    f"{billed_to.isoformat()} — the billed period was used")

            contract = contract_service.create_contract(
                client_id=client.id, property_id=prop.id,
                unit_ids=[u.id for u in take],
                start_date=start, expiry_date=max(expiry, start),
                monthly_rent=opening_rent,
                payment_mode=tenant["payment_mode"],
                opening_balance=tenant["opening_balance"] or 0,
                remarks=f"Migrated from New-2026 row {tenant['row']}",
                actor=actor)
            created["contracts"] += 1

            # --- the month columns, replayed as dated rent changes
            previous = opening_rent
            for month, value in timeline:
                if value is None or value == previous:
                    continue
                when = _iso(month)
                if value == 0:
                    break                    # a drop to zero is the tenancy ending
                contract_service.change_rent(
                    contract, new_rent=value, effective_date=when,
                    reason="Migrated from the workbook", actor=actor)
                created["amendments"] += 1
                previous = value

            if tenant["is_cancelled"] or stops_early:
                effective = _month_end(last_paid) if last_paid else start
                contract_service.cancel_contract(
                    contract, effective_date=max(effective, start),
                    reason=("Cancelled in the workbook" if tenant["is_cancelled"]
                            else "No rent recorded after this month in the workbook"),
                    actor=actor)
                created["cancellations"] += 1

        # --- what was paid to the landlord, month by month
        for month, amount in (block["rent_paid"] or {}).items():
            if not amount:
                continue
            already = LandlordPayment.query.filter_by(
                landlord_id=landlord.id, property_id=prop.id,
                period_month=_iso(month)).first()
            if already is not None:
                continue
            expense_service.post_landlord_payment(
                landlord_id=landlord.id, property_id=prop.id,
                period_month=_iso(month), amount=amount, mode="cash",
                reference="Migrated", actor=actor)
            created["landlord_payments"] += 1

        # --- the property's running costs
        for category_code, by_month in (block["expenses"] or {}).items():
            category = categories.get(category_code)
            if category is None:
                notes.append(f"No expense category with code {category_code!r} — "
                             f"costs for {block['property_name']} were skipped")
                continue
            for month, amount in by_month.items():
                if not amount:
                    continue
                exists = Expense.query.filter_by(
                    category_id=category.id, property_id=prop.id,
                    period_month=_iso(month), amount=amount).first()
                if exists is not None:
                    continue
                expense_service.post_expense(
                    category_id=category.id, period_month=_iso(month),
                    amount=amount, property_id=prop.id,
                    remarks="Migrated from the workbook", actor=actor)
                created["expenses"] += 1

    # --- debts owed by tenants who have already gone
    for row in plan.get("receivables", []):
        client = get_client(row["name"])
        prop = None
        if row.get("property_title"):
            wanted = _property_name(row["property_title"])
            prop = Property.query.filter(
                db.func.lower(Property.name) == wanted.lower()).first()
        if prop is None:
            notes.append(f"{row['name']}: receivable of {row['opening_balance']} "
                         "could not be tied to a property and was skipped")
            continue
        # These tenants had already left when the workbook was handed
        # over, so their tenancy is dated to the month *before* the
        # migrated year. That keeps them out of the way of the live
        # contracts competing for the same rooms, and makes the debt age
        # as what it is: older than everything else on the books.
        start = date(2025, 12, 1)
        finish = date(2025, 12, 31)
        if ClientContract.query.filter_by(
                client_id=client.id, property_id=prop.id, start_date=start).first():
            continue
        unit = Unit.query.filter_by(property_id=prop.id).order_by(Unit.id).first()
        if unit is None:
            notes.append(f"{row['name']}: receivable of {row['opening_balance']} "
                         f"skipped — {prop.name} has no units")
            continue
        contract = contract_service.create_contract(
            client_id=client.id, property_id=prop.id, unit_ids=[unit.id],
            start_date=start, expiry_date=finish, monthly_rent=0,
            payment_mode=row["payment_mode"],
            opening_balance=row["opening_balance"],
            remarks=f"Cancelled receivable, New-2026 row {row['row']}",
            actor=actor)
        contract_service.cancel_contract(
            contract, effective_date=finish,
            reason="Already cancelled when migrated", actor=actor)
        created["receivables"] += 1

    # --- raise the rent schedule so the months exist to compare against
    generated = rent_service.generate_all(
        upto=_iso(plan["months"][-1]) if plan.get("months") else None, actor=actor)

    return {"created": created, "reused": reused, "notes": notes,
            "rent": generated}


# ======================================================================
# The parallel run
# ======================================================================

TOLERANCE = 1.00       # Qatari Riyals — a rounding difference, not a fault


def reconcile(parsed: dict, month: date | None = None,
              plan: dict | None = None) -> dict:
    """Compare the app's figures for one month against the workbook's own.

    This is what makes the migration trustworthy: not "the importer said
    OK", but "the portal and the spreadsheet, asked the same question,
    give the same answer". Every line is shown with both figures and the
    difference — a reader can see *where* they part company, which a
    pass/fail could never tell them.
    """
    from .expenses import property_pnl
    from .receipts import collections_summary
    from .rent import month_start

    months = parsed.get("months") or []
    if month is None:
        month = _iso(months[-2]) if len(months) > 1 else (_iso(months[0]) if months else date.today())
    period = month_start(month)
    key = period.isoformat()

    totals = parsed.get("company_totals", {})
    pnl = property_pnl(period_month=period)
    collections = collections_summary(month=period)
    app_totals = pnl.get("totals", {})

    def line(label, workbook, app, note=None):
        workbook = None if workbook is None else round(float(workbook), 2)
        app = None if app is None else round(float(app), 2)
        if workbook is None or app is None:
            difference, matches = None, None
        else:
            difference = round(app - workbook, 2)
            matches = abs(difference) <= TOLERANCE
        return {"label": label, "workbook": workbook, "app": app,
                "difference": difference, "matches": matches, "note": note}

    def from_workbook(total_key: str, block_key: str,
                      category: str | None = None) -> float | None:
        """The sheet's own figure for this line.

        Prefer the roll-up at the foot of the sheet, because that is what
        the owner reads. Where a workbook has no roll-up, add the blocks
        up instead rather than reporting nothing — a line that cannot be
        compared teaches the operator less than one that can.
        """
        stated = (totals.get(total_key) or {}).get(key)
        if stated is not None:
            return stated
        running = 0.0
        found = False
        for block in parsed["blocks"]:
            if category:
                value = ((block.get("expenses") or {}).get(category) or {}).get(key)
            else:
                value = (block.get(block_key) or {}).get(key)
            if value is not None:
                running += value
                found = True
        return round(running, 2) if found else None

    lines = [
        line("Rent charged to clients",
             from_workbook("rent_received", "total"),
             collections.get("charged"),
             "The sheet's monthly column is what was billed for the month."),
        line("Paid to landlords",
             from_workbook("rent_paid", "rent_paid"),
             app_totals.get("rent_paid")),
        line("Sewage removal and cleaning",
             from_workbook("sewage", "expenses", "SEWAGE"),
             _category_total(pnl, "SEWAGE")),
        line("Electricity and water",
             from_workbook("electricity_water", "expenses", "ELECTRICITY_CAMP"),
             _category_total(pnl, "ELECTRICITY_CAMP")),
        line("Maintenance and cleaning",
             from_workbook("maintenance", "expenses", "MAINT_CLEANING"),
             _category_total(pnl, "MAINT_CLEANING")),
    ]

    # Per-property margin, which is the number the whole system exists
    # for. Names come from the plan when there is one, because that is
    # where two blocks on the same street were told apart — re-deriving
    # here would compare both of them against the same property.
    plan = plan or build_commit_plan(parsed)
    by_name = {r["property_name"]: r for r in pnl.get("rows", [])}
    profit_by_index = {b["index"]: (b.get("profit") or {}) for b in parsed["blocks"]}
    property_lines = []
    for block in plan["blocks"]:
        name = block["property_name"]
        workbook_profit = profit_by_index.get(block["index"], {}).get(key)
        row = by_name.get(name)
        property_lines.append(line(name, workbook_profit,
                                   row["profit"] if row else None))

    checked = [l for l in lines + property_lines if l["matches"] is not None]
    return {
        "month": key,
        "tolerance": TOLERANCE,
        "company": lines,
        "properties": property_lines,
        "summary": {
            "checked": len(checked),
            "matching": sum(1 for l in checked if l["matches"]),
            "differing": sum(1 for l in checked if not l["matches"]),
            "uncomparable": sum(1 for l in lines + property_lines
                                if l["matches"] is None),
        },
    }


def _category_total(pnl: dict, code: str) -> float:
    """Sum one expense category across every property.

    The P&L keys its expense breakdown by category *name* — which the
    operator can re-word in Settings — so resolve the code to whatever
    that category is currently called before looking it up.
    """
    from ..models import ExpenseCategory
    category = ExpenseCategory.query.filter_by(code=code).first()
    if category is None:
        return 0.0
    return round(sum((row.get("expenses") or {}).get(category.name, 0)
                     for row in pnl.get("rows", [])), 2)
